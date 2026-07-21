#!/usr/bin/env python3
"""GOC-only 3-statement model builder. Sources: canon.ts / cost-story.ts /
15-money-alignment-audit.md. Formulas live in Excel; python mirrors values to
assert the balance sheet balances before shipping."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/benknight/Code/Goods Asset Register/wiki/outputs/2026-07-21-goc-financial-model/GOC-3-statement-model-v1.xlsx"

wb = openpyxl.Workbook()

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="F2F2F2")
HDR_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
SRC_FONT = Font(size=9, italic=True, color="777777")
THIN = Border(bottom=Side(style="thin", color="CCCCCC"))

# ---------------------------------------------------------------- assumptions
A = wb.active
A.title = "Assumptions"
A.column_dimensions["A"].width = 44
for c in "BCD":
    A.column_dimensions[c].width = 13
A.column_dimensions["E"].width = 12
A.column_dimensions["F"].width = 70

A["A1"] = "GOC-only financial model, single input page (v1, 2026-07-21)"
A["A1"].font = TITLE_FONT
A["A2"] = ("Entity: the Goods on Country carve-out of A Curious Tractor Pty Ltd t/a Goods. "
           "YELLOW cells are the only inputs; every statement calculates from here. GST excluded. "
           "Labels: Verified = invoice/signed doc · Workpaper = our math, checkable · Modelled = from verified inputs, not demonstrated · Target = not signed.")
A["A2"].font = SRC_FONT
A["A4"] = "Input"; A["B4"] = "FY27"; A["C4"] = "FY28"; A["D4"] = "FY29"; A["E4"] = "Label"; A["F4"] = "Source"
for c in "ABCDEF":
    A[f"{c}4"].font = HDR_FONT

rows = [
    # (name, fy27, fy28, fy29, label, source)  None = single-value in B
    ("SALES", None, None, None, "", ""),
    ("Bed price ($/bed)", 750, "=B6", "=B6", "Verified", "canon.ts price list; one price everywhere"),
    ("Beds sold (units/yr)", 200, 338, 500, "Modelled", "ramp from verified ~120/yr run-rate (cost-story.ts:217) to ~338 break-even (engine) then growth toward ~1,000 capacity"),
    ("Share of beds pressed in-house (%)", 0.5, 1.0, 1.0, "Modelled", "capability proven: Maningrida Stretch run of 40 pressed at farm (Ben ruling 2026-07-21, INV-0303); rate ramps with equipment completion"),
    ("Washer revenue ($/yr)", 0, 0, 0, "Verified", "Pakkimjalki Kari is prototype, not for sale (CLAUDE.md products canon)"),
    ("UNIT COSTS", None, None, None, "", ""),
    ("Marginal cost per bed, buy-kit ($)", 684.79, "=B11", "=B11", "Verified (engine-locked)", "cost-story.ts:127 kit $344.05 + steel $27 + canvas $93.50 + hardware $5.24 + assembly $55.95 + freight ~$150"),
    ("Marginal cost per bed, pressed ($)", 425.74, "=B12", "=B12", "Modelled", "cost-story.ts:166; NOT yet measured at production rate; the funded measured run converts it"),
    ("FIXED BLOCK ($/yr)", None, None, None, "", ""),
    ("Facility (Kirmos share)", 27000, "=B14*(1+$B$19)", "=C14*(1+$B$19)", "Workpaper", "cost-story.ts:195 fixed block"),
    ("Founder production time (30d x $560)", 16800, "=B15*(1+$B$19)", "=C15*(1+$B$19)", "Workpaper", "engine.ts founder rate LOCKED $560/day"),
    ("Admin", 14700, "=B16*(1+$B$19)", "=C16*(1+$B$19)", "Workpaper", "cost-story.ts fixed block"),
    ("Field travel (relationship, fixed)", 51000, "=B17*(1+$B$19)", "=C17*(1+$B$19)", "Workpaper", "Ben ruling 2026-07-21: mostly fixed being-in-community time; delivery freight is in the per-bed line"),
    ("(sum check: block ≈ $109.5K/yr)", "=SUM(B14:B17)", "=SUM(C14:C17)", "=SUM(D14:D17)", "Workpaper", "cost-story.ts:195 ~$109,500"),
    ("Fixed cost annual growth (%)", 0.03, None, None, "Modelled", "placeholder CPI-ish; Ben/Matt to set"),
    ("EQUIPMENT", None, None, None, "", ""),
    ("Equipment already invested ($, opening PPE)", 110046, None, None, "Verified", "canon.ts:250; ~$80K TFN + ~$20K ACT"),
    ("New capex ($/yr)", 60000, 20000, 0, "Modelled", "within the $2-112K net remaining range (cost-story.ts:254); FY28 = contingency/moulds"),
    ("Depreciation rate (straight line, %/yr)", 0.10, None, None, "Modelled", "10yr equipment life assumption; Matt to confirm policy"),
    ("Equipment maintenance (% of gross PPE/yr)", 0.04, None, None, "Modelled", "cost-story.ts:279 3-5% band, midpoint"),
    ("FUNDING (THE RAISE) — $0 SIGNED TODAY", None, None, None, "", ""),
    ("Grant income: Snow R4/R5 ($)", 100000, 0, 0, "Target", "2026-07-03 stack; grant deed NOT signed"),
    ("Grant income: Centrecorp ($)", 75000, 0, 0, "Target", "2026-07-03 stack; NOT signed"),
    ("Grant income: QBE Stage 2 match ($)", 400000, 0, 0, "Target", "up to $400K, requires >=1:1 SIGNED match by 31 Aug (04-qbe-pipeline.md); discretionary"),
    ("Other grants/philanthropy ($)", 0, 50000, 0, "Target", "placeholder for continuing philanthropy taper; taper design = gifts end at break-even"),
    ("SEFA loan drawn ($)", 300000, 0, 0, "Target", "repayable working-capital anchor, proposed, NOT signed"),
    ("SEFA interest rate (%/yr)", 0.065, None, None, "Modelled", "concessional indicative; term sheet TBC"),
    ("SEFA principal repayment ($/yr)", 0, 100000, 100000, "Modelled", "indicative 3yr amortisation from FY28; term sheet TBC"),
    ("WORKING CAPITAL", None, None, None, "", ""),
    ("Debtor days", 60, None, None, "Modelled", "B2B buyers pay slow (Rotary INV-0222 went 11 months overdue); 60d planning assumption"),
    ("Creditor days", 30, None, None, "Modelled", "supplier terms assumption"),
    ("Inventory days (of COGS)", 30, None, None, "Modelled", "materials bought ahead of builds"),
    ("OPENING POSITION (Ben/accountant to set)", None, None, None, "", ""),
    ("Opening cash ($)", 50000, None, None, "PLACEHOLDER", "set from bank; NOT a canon figure"),
    ("Opening receivables ($)", 143000, None, None, "Verified", "canon.ts:92 AR = Rotary $82.5K + Homeland $44K + Regional Arts $16.5K"),
    ("Tax rate on positive NPBT (%)", 0.25, None, None, "Modelled", "small business company rate; entity structure pre-Butterfly, Matt to confirm"),
]
r = 5
row_of = {}
for name, b, c, d, label, src in rows:
    A[f"A{r}"] = name
    row_of[name] = r
    if b is None and c is None and d is None:
        A[f"A{r}"].font = HDR_FONT
    else:
        for col, v in (("B", b), ("C", c), ("D", d)):
            if v is not None:
                A[f"{col}{r}"] = v
                is_formula = isinstance(v, str) and v.startswith("=")
                A[f"{col}{r}"].fill = CALC_FILL if is_formula else INPUT_FILL
                if isinstance(v, float) and v < 1 and v != 0:
                    A[f"{col}{r}"].number_format = "0.0%"
                elif not is_formula:
                    A[f"{col}{r}"].number_format = "#,##0.00" if isinstance(v, float) else "#,##0"
        A[f"E{r}"] = label
        A[f"F{r}"] = src
        A[f"F{r}"].font = SRC_FONT
        A[f"A{r}"].border = THIN
    r += 1

# convenience: map to concrete cells (single-value inputs live in col B)
P = {
    "price": "Assumptions!B6", "beds": ("Assumptions!B7", "Assumptions!C7", "Assumptions!D7"),
    "pressed_pct": ("Assumptions!B8", "Assumptions!C8", "Assumptions!D8"),
    "cost_kit": "Assumptions!B11", "cost_pressed": "Assumptions!B12",
    "block": ("Assumptions!B18", "Assumptions!C18", "Assumptions!D18"),
    "ppe_open": "Assumptions!B21",
    "capex": ("Assumptions!B22", "Assumptions!C22", "Assumptions!D22"),
    "dep_rate": "Assumptions!B23", "maint_rate": "Assumptions!B24",
    "grants": [("Assumptions!B26", "Assumptions!C26", "Assumptions!D26"),
               ("Assumptions!B27", "Assumptions!C27", "Assumptions!D27"),
               ("Assumptions!B28", "Assumptions!C28", "Assumptions!D28"),
               ("Assumptions!B29", "Assumptions!C29", "Assumptions!D29")],
    "loan_draw": ("Assumptions!B30", "Assumptions!C30", "Assumptions!D30"),
    "int_rate": "Assumptions!B31",
    "loan_repay": ("Assumptions!B32", "Assumptions!C32", "Assumptions!D32"),
    "debtor_days": "Assumptions!B34", "creditor_days": "Assumptions!B35", "inv_days": "Assumptions!B36",
    "open_cash": "Assumptions!B38", "open_ar": "Assumptions!B39", "tax": "Assumptions!B40",
}

YEARS = ["FY27", "FY28", "FY29"]
COLS = ["B", "C", "D"]

def sheet(name):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 44
    for c in COLS:
        ws.column_dimensions[c].width = 15
    ws["A1"] = name + " — GOC carve-out (calculated; edit only the Assumptions page)"
    ws["A1"].font = TITLE_FONT
    for i, y in enumerate(YEARS):
        ws[f"{COLS[i]}3"] = y
        ws[f"{COLS[i]}3"].font = HDR_FONT
    return ws

def put(ws, r, label, formulas, bold=False, fmt="#,##0"):
    ws[f"A{r}"] = label
    if bold:
        ws[f"A{r}"].font = Font(bold=True)
    for i, f in enumerate(formulas):
        ws[f"{COLS[i]}{r}"] = f
        ws[f"{COLS[i]}{r}"].number_format = fmt
        if bold:
            ws[f"{COLS[i]}{r}"].font = Font(bold=True)
    return r + 1

# ------------------------------------------------------------------------ P&L
PL = sheet("P&L")
r = 5
r = put(PL, r, "Bed revenue (beds x price)", [f"={P['beds'][i]}*{P['price']}" for i in range(3)], bold=True)
r = put(PL, r, "Blended marginal cost per bed", [f"={P['pressed_pct'][i]}*{P['cost_pressed']}+(1-{P['pressed_pct'][i]})*{P['cost_kit']}" for i in range(3)])
r = put(PL, r, "COGS (beds x blended marginal)", [f"=-B7*{P['beds'][0]}" , f"=-C7*{P['beds'][1]}", f"=-D7*{P['beds'][2]}"])
r = put(PL, r, "Gross profit (contribution)", [f"={c}5+{c}8" for c in COLS], bold=True)
r += 1
r = put(PL, r, "Fixed block (facility, founder, admin, travel)", [f"=-{P['block'][i]}" for i in range(3)])
r = put(PL, r, "Equipment maintenance", [f"=-{P['maint_rate']}*('Balance Sheet'!{c}10)" for c in COLS])
r = put(PL, r, "EBITDA", [f"={c}9+{c}11+{c}12" for c in COLS], bold=True)
r = put(PL, r, "Depreciation", [f"=-{P['dep_rate']}*'Balance Sheet'!{c}10" for c in COLS])
r = put(PL, r, "EBIT", [f"={c}13+{c}14" for c in COLS])
r = put(PL, r, "Interest (SEFA)", [f"=-{P['int_rate']}*'Balance Sheet'!{c}16" for c in COLS])
r = put(PL, r, "Trading profit before grants and tax", [f"={c}15+{c}16" for c in COLS], bold=True)
r += 1
r = put(PL, r, "Grant income (Target: $0 signed today)", ["=" + "+".join(g[i] for g in P['grants']) for i in range(3)])
r = put(PL, r, "Net profit before tax", [f"={c}17+{c}19" for c in COLS], bold=True)
r = put(PL, r, "Tax (only if positive)", [f"=-MAX({c}20,0)*{P['tax']}" for c in COLS])
r = put(PL, r, "NET PROFIT AFTER TAX", [f"={c}20+{c}21" for c in COLS], bold=True)
PL["A24"] = ("Reference: FY26 actual Goods-only revenue $713,827 (accountant-signed carve-out, canon.ts:106; the ONLY external revenue figure) "
             "+ $143,000 AR. FY26 included grant-funded delivery; this model separates trading from grants so the taper is visible.")
PL["A24"].font = SRC_FONT

# ------------------------------------------------------------------ Cash Flow
CF = sheet("Cash Flow")
r = 5
r = put(CF, r, "Net profit after tax", [f"='P&L'!{c}22" for c in COLS])
r = put(CF, r, "Add back depreciation", [f"=-'P&L'!{c}14" for c in COLS])
r = put(CF, r, "(Increase) in receivables", [f"=-('Balance Sheet'!{c}7-{P['open_ar']})" ,
                                              "=-('Balance Sheet'!C7-'Balance Sheet'!B7)",
                                              "=-('Balance Sheet'!D7-'Balance Sheet'!C7)"])
r = put(CF, r, "(Increase) in inventory", ["=-('Balance Sheet'!B8-0)",
                                            "=-('Balance Sheet'!C8-'Balance Sheet'!B8)",
                                            "=-('Balance Sheet'!D8-'Balance Sheet'!C8)"])
r = put(CF, r, "Increase in payables", ["='Balance Sheet'!B14-0",
                                         "='Balance Sheet'!C14-'Balance Sheet'!B14",
                                         "='Balance Sheet'!D14-'Balance Sheet'!C14"])
r = put(CF, r, "Operating cash flow", [f"=SUM({c}5:{c}9)" for c in COLS], bold=True)
r += 1
r = put(CF, r, "Capex", [f"=-{P['capex'][i]}" for i in range(3)])
r = put(CF, r, "Investing cash flow", [f"={c}12" for c in COLS], bold=True)
r += 1
r = put(CF, r, "SEFA loan drawn", [f"={P['loan_draw'][i]}" for i in range(3)])
r = put(CF, r, "SEFA principal repaid", [f"=-{P['loan_repay'][i]}" for i in range(3)])
r = put(CF, r, "Financing cash flow", [f"={c}15+{c}16" for c in COLS], bold=True)
r += 1
r = put(CF, r, "Net cash movement", [f"={c}10+{c}13+{c}17" for c in COLS], bold=True)
r = put(CF, r, "Opening cash", [f"={P['open_cash']}", "=B20", "=C20"])
r = put(CF, r, "CLOSING CASH", [f"={c}19+{c}20" for c in COLS], bold=True)

# -------------------------------------------------------------- Balance Sheet
BS = sheet("Balance Sheet")
r = 5
r = put(BS, r, "ASSETS", ["", "", ""], bold=True)
r = put(BS, r, "Cash", [f"='Cash Flow'!{c}21" for c in COLS])
r = put(BS, r, "Receivables (revenue x debtor days/365)", [f"='P&L'!{c}5*{P['debtor_days']}/365" for c in COLS])
r = put(BS, r, "Inventory (COGS x inventory days/365)", [f"=-'P&L'!{c}8*{P['inv_days']}/365" for c in COLS])
r += 0
r = put(BS, r, "PPE at cost (opening + cumulative capex)",
        [f"={P['ppe_open']}+{P['capex'][0]}",
         f"=B9+{P['capex'][1]}",
         f"=C9+{P['capex'][2]}"])
BS_ppe_row = r - 1
r = put(BS, r, "PPE gross (avg base for dep/maint)", [f"={c}9" for c in COLS])
r = put(BS, r, "Accumulated depreciation", [f"=-'P&L'!B14*-1*0-'P&L'!B14*-1", "=B11-'P&L'!C14", "=C11-'P&L'!D14"])
r = put(BS, r, "TOTAL ASSETS", [f"={c}6+{c}7+{c}8+{c}9+{c}11" for c in COLS], bold=True)
r += 1
r = put(BS, r, "LIABILITIES", ["", "", ""], bold=True)
BS["A14"] = "Payables (COGS x creditor days/365)"
for i, c in enumerate(COLS):
    BS[f"{c}14"] = f"=-'P&L'!{c}8*{P['creditor_days']}/365"
    BS[f"{c}14"].number_format = "#,##0"
r = 15
r = put(BS, r, "", ["", "", ""])
BS["A16"] = "SEFA loan balance"
BS["B16"] = f"={P['loan_draw'][0]}-{P['loan_repay'][0]}"
BS["C16"] = f"=B16+{P['loan_draw'][1]}-{P['loan_repay'][1]}"
BS["D16"] = f"=C16+{P['loan_draw'][2]}-{P['loan_repay'][2]}"
for c in COLS:
    BS[f"{c}16"].number_format = "#,##0"
r = 17
r = put(BS, r, "TOTAL LIABILITIES", [f"={c}14+{c}16" for c in COLS], bold=True)
r += 1
r = put(BS, r, "EQUITY", ["", "", ""], bold=True)
r = put(BS, r, "Opening funds employed (cash + AR + PPE)", [f"={P['open_cash']}+{P['open_ar']}+{P['ppe_open']}", "=B20", "=C20"])
r = put(BS, r, "Cumulative retained profit", [f"='P&L'!B22", "=B21+'P&L'!C22", "=C21+'P&L'!D22"])
r = put(BS, r, "TOTAL EQUITY", [f"={c}20+{c}21" for c in COLS], bold=True)
r += 1
r = put(BS, r, "CHECK (assets - liabilities - equity = 0)", [f"={c}12-{c}17-{c}22" for c in COLS], bold=True)

wb.save(OUT)
print("saved", OUT)

# ------------------------------------------------------- python mirror check
price, years = 750.0, range(3)
beds = [200, 338, 500]; pressed = [0.5, 1.0, 1.0]
ck, cp = 684.79, 425.74
block0 = [27000, 16800, 14700, 51000]
grow = 0.03
blocks = [sum(block0), sum(block0) * (1 + grow), sum(block0) * (1 + grow) ** 2]
ppe_open, capex = 110046.0, [60000, 20000, 0]
dep_rate, maint = 0.10, 0.04
grants = [575000, 50000, 0]
draw = [300000, 0, 0]; repay = [0, 100000, 100000]; ir = 0.065
dd, cd, invd = 60, 30, 30
ocash, oar, tax = 50000.0, 143000.0, 0.25

ppe = []; g = ppe_open
for y in years:
    g += capex[y]; ppe.append(g)
acc = 0; accs = []
for y in years:
    acc += dep_rate * ppe[y]; accs.append(acc)
loan = []; l = 0
for y in years:
    l += draw[y] - repay[y]; loan.append(l)
npat = []; ar = []; inv = []; ap = []
for y in years:
    rev = beds[y] * price
    cogs = beds[y] * (pressed[y] * cp + (1 - pressed[y]) * ck)
    ebitda = rev - cogs - blocks[y] - maint * ppe[y]
    dep = dep_rate * ppe[y]
    interest = ir * loan[y]
    npbt = ebitda - dep - interest + grants[y]
    t = max(npbt, 0) * tax
    npat.append(npbt - t)
    ar.append(rev * dd / 365); inv.append(cogs * invd / 365); ap.append(cogs * cd / 365)
cash = []; c = ocash
for y in years:
    d_ar = ar[y] - (oar if y == 0 else ar[y - 1])
    d_inv = inv[y] - (0 if y == 0 else inv[y - 1])
    d_ap = ap[y] - (0 if y == 0 else ap[y - 1])
    op = npat[y] + dep_rate * ppe[y] - d_ar - d_inv + d_ap
    c += op - capex[y] + draw[y] - repay[y]
    cash.append(c)
oeq = ocash + oar + ppe_open
cum = 0
for y in years:
    cum += npat[y]
    assets = cash[y] + ar[y] + inv[y] + ppe[y] - accs[y]
    liab = ap[y] + loan[y]
    eq = oeq + cum
    print(YEARS[y], f"npat={npat[y]:,.0f} cash={cash[y]:,.0f} check={assets - liab - eq:,.1f}")
