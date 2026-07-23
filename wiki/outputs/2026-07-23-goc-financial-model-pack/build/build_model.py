#!/usr/bin/env python3
"""Build the GOC-only 3-statement financial model (live formulas).
Sheets: Assumptions (single input page) -> Production -> P&L -> BalanceSheet
-> CashFlow -> Sources. Mirrors shadow_calc.py exactly. All AUD ex-GST.
Blue = input (override-able). Black = formula."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT="/Users/benknight/Code/Goods Asset Register/wiki/outputs/2026-07-23-goc-financial-model-pack/GOC-3-Statement-Model.xlsx"
wb=openpyxl.Workbook()

BLUE=Font(color="0000CC")               # input
HDR=Font(bold=True, size=12)
SEC=Font(bold=True, color="FFFFFF")
SECFILL=PatternFill("solid", fgColor="44546A")
TOT=Font(bold=True)
GREY=Font(color="808080", italic=True, size=9)
CUR='#,##0;(#,##0)'; PC='0%'; TWO='0.00'; DAY='0'
YC={1:'C',2:'D',3:'E',4:'F',5:'G'}      # year columns
YEARS=[1,2,3,4,5]

def sname(s): return s
def ref(sheet,cell): return f"'{sheet}'!{cell}"

# ============================ ASSUMPTIONS ============================
A=wb.active; A.title="Assumptions"
A.column_dimensions['A'].width=42
for c in ['B','C','D','E','F','G']: A.column_dimensions[c].width=13
A.column_dimensions['H'].width=52
A['A1']="Goods on Country — GOC-only 3-Statement Model"; A['A1'].font=HDR
A['A2']=("Blue = input (override-able). Black = formula. All AUD ex-GST. 5-year horizon (FY1=FY2026-27). "
         "Defaults = recommended pending Ben's six decisions (see Assumptions notes + the pack's alignment doc). "
         "Base case = the central-production investment case (Pot 1); the community wraparound (Pot 2) is fully wired but zeroed until the rollout is set.")
A['A2'].font=GREY; A['A2'].alignment=Alignment(wrap_text=True)
A.row_dimensions[2].height=42

def sec(r,label):
    A[f'A{r}']=label; A[f'A{r}'].font=SEC; A[f'A{r}'].fill=SECFILL
    for c in ['B','C','D','E','F','G','H']:
        A[f'{c}{r}'].fill=SECFILL
def inp(r,label,val,note="",fmt=CUR):     # single-value input in col B
    A[f'A{r}']=label; A[f'B{r}']=val; A[f'B{r}'].font=BLUE; A[f'B{r}'].number_format=fmt
    if note: A[f'H{r}']=note; A[f'H{r}'].font=GREY; A[f'H{r}'].alignment=Alignment(wrap_text=True)
    return f'$B${r}'
def frm(r,label,formula,note="",fmt=CUR):  # single-value formula in col B
    A[f'A{r}']=label; A[f'B{r}']=formula; A[f'B{r}'].font=TOT; A[f'B{r}'].number_format=fmt
    if note: A[f'H{r}']=note; A[f'H{r}'].font=GREY
    return f'$B${r}'
def yhdr(r):
    for y in YEARS:
        A[f'{YC[y]}{r}']=f"Year {y}"; A[f'{YC[y]}{r}'].font=Font(bold=True)
def sched(r,label,vals,note="",fmt=CUR):   # year-schedule input across C..G
    A[f'A{r}']=label
    for y in YEARS:
        A[f'{YC[y]}{r}']=vals[y-1]; A[f'{YC[y]}{r}'].font=BLUE; A[f'{YC[y]}{r}'].number_format=fmt
    if note: A[f'H{r}']=note; A[f'H{r}'].font=GREY; A[f'H{r}'].alignment=Alignment(wrap_text=True)
    return {y:f'{YC[y]}{r}' for y in YEARS}

sec(4,"1. PRICE & UNIT ECONOMICS  (Pot 1 production — from the bed model)")
a_price=inp(5,"Selling price per bed",750,"Verified: shop stretch-bed-single, held flat.",CUR)
a_kit=inp(6,"Buy-Kit marginal cost per bed",684.79,"Modelled (verified inputs). Today's bought-in path.",TWO)
a_fac=inp(7,"Factory marginal cost per bed",425.74,"PROCESS PROVEN (40-bed Maningrida in-house run); per-bed cost modelled pending measured actuals.",TWO)
a_com=inp(8,"Community marginal cost per bed",420.74,"Modelled; on-country community path is future.",TWO)

sec(10,"2. PRODUCTION DRIVERS")
a_factories=inp(11,"Number of factories",1,"Sunshine Coast facility, constant.",DAY)
a_bedsfac=inp(12,"Beds per factory per year",500,"Practical max ~1,000/yr.",DAY)
a_bedscont=inp(13,"Beds per container per year",500,"Same ~1,000/yr practical max. OPEN: 250 vs 1,250 vs 1,500 — a measured run confirms.",DAY)
a_kitmax=inp(14,"Kit (Defy) beds per year (max)",100,"Auto-tapers to 0 once in-house covers it.",DAY)
a_util=inp(15,"First-year utilisation of new facilities",0.40,"Ramp: facilities run at 40% in their build year, 100% after.",PC)
yhdr(16)
a_newcont=sched(17,"New containers built per year",[1,0,0,1,2],"Explicit build plan (Ben decision). Matches Matt's endogenous ramp Y1-5.",DAY)

sec(19,"3. FIXED OPERATING BLOCK  (Pot 1, per year)")
a_founder=inp(20,"Founder production time (30d x $560)",16800,"Locked 2026-05-29. Only production days touch unit cost.")
a_kirmos=inp(21,"Kirmos facility (50% to beds)",27000,"GoC Q&A Q1 confirms.")
a_admin=inp(22,"Admin",14700,"")
a_travel=inp(23,"Field travel",51000,"")
a_fixed=frm(24,"TOTAL fixed block / yr","=SUM(B20:B23)","= $109,500 (confirmed line by line).")
a_equipsub=inp(25,"Equipment subtotal (maintenance base)",167000,"Shredder+press+CNC+bench midpoints.")
a_maint=inp(26,"Equipment maintenance % of subtotal",0.05,"Placeholder 3-5% pending vendor quotes; per facility/yr.",PC)

sec(28,"4. CAPEX & DEPRECIATION")
a_facbuild=inp(29,"Facility build cost per container",207450,"Matt Inputs B27 (equipment $167K + container fit-out $40,450). Capitalised.")
a_openplant=inp(30,"Opening plant (sunk, evidenced)",75000,"MVF evidenced sunk hardware (Nic, used deals). Ownership handover targeted end-Aug 2026.")
a_life=inp(31,"Plant useful life (years)",10,"Straight-line, full year in year of acquisition (simplification).",DAY)

sec(33,"5. WORKING CAPITAL & TAX")
a_debtor=inp(34,"Debtor days",30,"On product revenue.",DAY)
a_creditor=inp(35,"Creditor days",30,"On cash operating costs.",DAY)
a_invdays=inp(36,"Inventory days",45,"On bed COGS.",DAY)
a_tax=inp(37,"Company tax rate",0.25,"Pty Ltd base rate. Confirm charity/DGR treatment with accountant. Loss carry-forward modelled.",PC)
a_int=inp(38,"Interest rate on debt",0.06,"On QBE+SEFA borrowings, interest-only in the 5-yr window.",PC)
a_opencash=inp(39,"Opening cash",50000,"Most important single input (base placeholder).")

sec(41,"6. CAPITAL STACK  (draw schedule)"); yhdr(41)
a_debtdraw=sched(42,"Debt draw (SEFA Yr1, QBE Yr2)",[300000,400000,0,0,0],"Repayable capital. PLACEHOLDER — confirm split + QBE match ratio. $0 signed today.")
a_equitydraw=sched(43,"Equity draw (anchor philanthropy)",[500000,0,0,0,0],"Contributed capital / grant to capital. PLACEHOLDER.")

sec(45,"7. POT 2 — COMMUNITY WRAPAROUND  (grant-funded; ZERO in base case)"); yhdr(45)
a_active=sched(46,"Active community sites",[0,0,0,0,0],"Base case = 0. Set >0 to bring the community rollout (Pot 2) into the statements.",DAY)
a_bare=inp(47,"Bare facility op cost / site / yr",151666,"DEWR-derived: rent+insurance+admin+upkeep. Consortium/grant-carried.")
a_staffup=inp(48,"Staffed uplift / site / yr",190000,"Manager $150K + trainer/WHS ~$40K.")
a_program=inp(49,"Employment program / yr (if active)",300000,"Brokerage/wages. Grant money, never bed sales.")
a_grant=sched(50,"Grant income (Pot 2)",[0,0,0,0,0],"Program/wraparound funding, matched to Pot 2 costs. Base = 0.")

a_opencapital=frm(52,"Opening capital (sunk plant + opening cash)","=B30+B39","Opening equity so the Year-0 balance sheet balances.")

# helper to build Assumptions absolute ref
def AB(addr): return ref("Assumptions",addr)              # single ($B$n)
def AY(scheddict,y): return ref("Assumptions",f"${scheddict[y][0]}${scheddict[y][1:]}") if False else ref("Assumptions",scheddict[y])

# ============================ PRODUCTION ============================
P=wb.create_sheet("Production")
P.column_dimensions['A'].width=34
for c in ['B','C','D','E','F','G']: P.column_dimensions[c].width=13
P.column_dimensions['H'].width=40
P['A1']="Production schedule"; P['A1'].font=HDR
P['A2']="Drives the three statements. Factory beds = factories x rate (x util in Yr1). Community beds = containers x rate (new containers at util in build year). Kit auto-tapers to 0 once in-house covers the kit input."
P['A2'].font=GREY; P['A2'].alignment=Alignment(wrap_text=True); P.row_dimensions[2].height=30
P['A3']="Line";
for y in YEARS: P[f'{YC[y]}3']=f"Year {y}"; P[f'{YC[y]}3'].font=Font(bold=True)
def prow(r,label,fmt=DAY): P[f'A{r}']=label; [P.__setitem__(f'{YC[y]}{r}', None) for y in YEARS];
rows_P={}
def setP(r,label,fn,fmt=DAY,bold=False,note=""):
    P[f'A{r}']=label
    for y in YEARS:
        P[f'{YC[y]}{r}']=fn(y); P[f'{YC[y]}{r}'].number_format=fmt
        if bold: P[f'{YC[y]}{r}'].font=TOT
    if note: P[f'H{r}']=note; P[f'H{r}'].font=GREY
    return r
# rows
r_fact=4;  setP(r_fact,"Factories",lambda y: f"={AB(a_factories)}")
r_new=5;   setP(r_new,"New containers built",lambda y: f"={AY(a_newcont,y)}")
r_cont=6
P['A6']="Containers operating (cumulative)"
P['C6']=f"=C{r_new}"; P['C6'].number_format=DAY
for y in [2,3,4,5]:
    P[f'{YC[y]}6']=f"={YC[y-1]}6+{YC[y]}{r_new}"; P[f'{YC[y]}6'].number_format=DAY
r_facb=7
P['A7']="Factory beds"
P['C7']=f"={AB(a_factories)}*{AB(a_bedsfac)}*{AB(a_util)}"; P['C7'].number_format=DAY
for y in [2,3,4,5]:
    P[f'{YC[y]}7']=f"={AB(a_factories)}*{AB(a_bedsfac)}"; P[f'{YC[y]}7'].number_format=DAY
r_comb=8;  setP(r_comb,"Community beds",lambda y: f"=({YC[y]}{r_cont}-{YC[y]}{r_new})*{AB(a_bedscont)}+{YC[y]}{r_new}*{AB(a_bedscont)}*{AB(a_util)}")
r_kitb=9;  setP(r_kitb,"Kit beds",lambda y: f"=IF({YC[y]}{r_facb}+{YC[y]}{r_comb}>={AB(a_kitmax)},0,{AB(a_kitmax)})")
r_tot=10;  setP(r_tot,"Total beds",lambda y: f"={YC[y]}{r_facb}+{YC[y]}{r_comb}+{YC[y]}{r_kitb}",DAY,True)
r_rev=11;  setP(r_rev,"Product revenue",lambda y: f"={YC[y]}{r_tot}*{AB(a_price)}",CUR,True)
r_cogs=12; setP(r_cogs,"Bed COGS",lambda y: f"={YC[y]}{r_kitb}*{AB(a_kit)}+{YC[y]}{r_facb}*{AB(a_fac)}+{YC[y]}{r_comb}*{AB(a_com)}",CUR)
r_gm=13;   setP(r_gm,"Gross margin (production)",lambda y: f"={YC[y]}{r_rev}-{YC[y]}{r_cogs}",CUR,True)
r_maint=14;setP(r_maint,"Equipment maintenance",lambda y: f"={AB(a_maint)}*{AB(a_equipsub)}*({YC[y]}{r_fact}+{YC[y]}{r_cont})",CUR)
PR=dict(rev=r_rev,cogs=r_cogs,maint=r_maint,new=r_new,cont=r_cont,fact=r_fact,tot=r_tot,gm=r_gm)

def PY(r,y): return ref("Production",f"{YC[y]}{r}")

# ============================ BALANCE SHEET (with workings) ============================
B=wb.create_sheet("BalanceSheet")
B.column_dimensions['A'].width=34
for c in ['B','C','D','E','F','G']: B.column_dimensions[c].width=13
B['A1']="Balance Sheet (AUD)"; B['A1'].font=HDR
B['B3']="Opening"; B['B3'].font=Font(bold=True)
for y in YEARS: B[f'{YC[y]}3']=f"Year {y}"; B[f'{YC[y]}3'].font=Font(bold=True)
# --- workings block (rows 30+) ---
wr=31
B[f'A30']="WORKINGS"; B['A30'].font=Font(bold=True, italic=True)
# capex
B[f'A{wr}']="Capex (container builds)"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={PY(PR['new'],y)}*{AB(a_facbuild)}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_capex=wr; wr+=1
# cumulative capex
B[f'A{wr}']="Cumulative capex"
B[f'C{wr}']=f"=C{r_capex}"
for y in [2,3,4,5]: B[f'{YC[y]}{wr}']=f"={YC[y-1]}{wr}+{YC[y]}{r_capex}"
r_cumcapex=wr; [B.__setitem__(f'{YC[y]}{wr}', B[f'{YC[y]}{wr}'].value) for y in YEARS]
for y in YEARS: B[f'{YC[y]}{r_cumcapex}'].number_format=CUR
wr+=1
# gross PPE
B[f'A{wr}']="Gross PP&E"; B[f'B{wr}']=f"={AB(a_openplant)}"; B[f'B{wr}'].number_format=CUR
for y in YEARS: B[f'{YC[y]}{wr}']=f"={AB(a_openplant)}+{YC[y]}{r_cumcapex}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_gross=wr; wr+=1
# depreciation
B[f'A{wr}']="Depreciation (year)"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={YC[y]}{r_gross}/{AB(a_life)}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_dep=wr; wr+=1
# accum deprec
B[f'A{wr}']="Accumulated depreciation"; B[f'B{wr}']=0; B[f'B{wr}'].number_format=CUR
B[f'C{wr}']=f"=B{wr}+C{r_dep}"
for y in [2,3,4,5]: B[f'{YC[y]}{wr}']=f"={YC[y-1]}{wr}+{YC[y]}{r_dep}"
for y in YEARS: B[f'{YC[y]}{wr}'].number_format=CUR
r_accum=wr; wr+=1
# net PPE
B[f'A{wr}']="Net PP&E"; B[f'B{wr}']=f"={AB(a_openplant)}"; B[f'B{wr}'].number_format=CUR
for y in YEARS: B[f'{YC[y]}{wr}']=f"={YC[y]}{r_gross}-{YC[y]}{r_accum}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_netppe=wr; wr+=1
# debt roll
B[f'A{wr}']="Debt draw"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={AY(a_debtdraw,y)}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_ddraw=wr; wr+=1
B[f'A{wr}']="Borrowings (close)"; B[f'B{wr}']=0; B[f'B{wr}'].number_format=CUR
B[f'C{wr}']=f"=B{wr}+C{r_ddraw}"
for y in [2,3,4,5]: B[f'{YC[y]}{wr}']=f"={YC[y-1]}{wr}+{YC[y]}{r_ddraw}"
for y in YEARS: B[f'{YC[y]}{wr}'].number_format=CUR
r_debt=wr; wr+=1
B[f'A{wr}']="Interest expense"
B[f'C{wr}']=f"=B{r_debt}*{AB(a_int)}"
for y in [2,3,4,5]: B[f'{YC[y]}{wr}']=f"={YC[y-1]}{r_debt}*{AB(a_int)}"
for y in YEARS: B[f'{YC[y]}{wr}'].number_format=CUR
r_int=wr; wr+=1
# working capital
B[f'A{wr}']="Accounts receivable"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={AB(a_debtor)}/365*{PY(PR['rev'],y)}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_ar=wr; wr+=1
B[f'A{wr}']="Inventory"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={AB(a_invdays)}/365*{PY(PR['cogs'],y)}"; B[f'{YC[y]}{wr}'].number_format=CUR
r_inv=wr; wr+=1
B[f'A{wr}']="Accounts payable"
for y in YEARS: B[f'{YC[y]}{wr}']=f"={AB(a_creditor)}/365*({PY(PR['cogs'],y)}+{AB(a_fixed)}+{PY(PR['maint'],y)})"; B[f'{YC[y]}{wr}'].number_format=CUR
r_ap=wr; wr+=1
BW=dict(capex=r_capex,gross=r_gross,dep=r_dep,accum=r_accum,netppe=r_netppe,ddraw=r_ddraw,debt=r_debt,intr=r_int,ar=r_ar,inv=r_inv,ap=r_ap)
def BY(r,y): return ref("BalanceSheet",f"{YC[y]}{r}")
def BYo(r): return ref("BalanceSheet",f"B{r}")

# P&L needs BW; build P&L now
# ============================ P&L ============================
L=wb.create_sheet("P&L")
L.column_dimensions['A'].width=40
for c in ['B','C','D','E','F','G']: L.column_dimensions[c].width=13
L['A1']="Profit & Loss (accrual, AUD)"; L['A1'].font=HDR
for y in YEARS: L[f'{YC[y]}3']=f"Year {y}"; L[f'{YC[y]}3'].font=Font(bold=True)
def Lrow(r,label,fn,fmt=CUR,bold=False,sec=False,note=""):
    L[f'A{r}']=label
    if sec: L[f'A{r}'].font=Font(bold=True, italic=True)
    if fn:
        for y in YEARS:
            L[f'{YC[y]}{r}']=fn(y); L[f'{YC[y]}{r}'].number_format=fmt
            if bold: L[f'{YC[y]}{r}'].font=TOT
    if note: L[f'A{r}']=label
    return r
Lrow(4,"POT 1 — PRODUCTION (the investment case)",None,sec=True)
r_prev=Lrow(5,"Product revenue",lambda y: f"={PY(PR['rev'],y)}")
r_pcogs=Lrow(6,"Bed COGS",lambda y: f"=-{PY(PR['cogs'],y)}")
r_pgp=Lrow(7,"Gross profit",lambda y: f"={YC[y]}5+{YC[y]}6",bold=True)
r_pfix=Lrow(8,"Central fixed operating block",lambda y: f"=-{AB(a_fixed)}")
r_pmaint=Lrow(9,"Equipment maintenance",lambda y: f"=-{PY(PR['maint'],y)}")
r_pebitda=Lrow(10,"Production EBITDA",lambda y: f"={YC[y]}7+{YC[y]}8+{YC[y]}9",bold=True)
Lrow(12,"POT 2 — WRAPAROUND (grant-funded; zero in base case)",None,sec=True)
r_grantinc=Lrow(13,"Grant / program income",lambda y: f"={AY(a_grant,y)}")
r_barec=Lrow(14,"Community-site facility operating",lambda y: f"=-{AY(a_active,y)}*{AB(a_bare)}")
r_progc=Lrow(15,"Program & staff costs",lambda y: f"=-({AY(a_active,y)}*{AB(a_staffup)}+IF({AY(a_active,y)}>0,{AB(a_program)},0))")
r_pot2=Lrow(16,"Pot 2 net (grant less program)",lambda y: f"={YC[y]}13+{YC[y]}14+{YC[y]}15",bold=True)
Lrow(18,"ENTITY",None,sec=True)
r_ebitda=Lrow(19,"EBITDA",lambda y: f"={YC[y]}{r_pebitda}+{YC[y]}{r_pot2}",bold=True)
r_dep=Lrow(20,"Depreciation",lambda y: f"=-{BY(BW['dep'],y)}")
r_ebit=Lrow(21,"EBIT",lambda y: f"={YC[y]}{r_ebitda}+{YC[y]}{r_dep}",bold=True)
r_intr=Lrow(22,"Interest expense",lambda y: f"=-{BY(BW['intr'],y)}")
r_npbt=Lrow(23,"NPBT",lambda y: f"={YC[y]}{r_ebit}+{YC[y]}{r_intr}",bold=True)
# tax with loss carryforward: working rows below
r_lossopen=27; r_taxable=28; r_lossclose=29
L['A26']="Tax working (loss carry-forward)"; L['A26'].font=Font(bold=True, italic=True)
L['A27']="Loss pool (opening)"; L['B27']=0; L['B27'].number_format=CUR
L['C27']="=B27"
for y in [2,3,4,5]: L[f'{YC[y]}27']=f"={YC[y-1]}{r_lossclose}"
for y in YEARS: L[f'{YC[y]}27'].number_format=CUR
L['A28']="Taxable profit"
for y in YEARS: L[f'{YC[y]}28']=f"=MAX(0,{YC[y]}{r_npbt}-{YC[y]}{r_lossopen})"; L[f'{YC[y]}28'].number_format=CUR
L['A29']="Loss pool (closing)"
for y in YEARS: L[f'{YC[y]}29']=f"=MAX(0,{YC[y]}{r_lossopen}-MAX(0,{YC[y]}{r_npbt}))+MAX(0,-{YC[y]}{r_npbt})"; L[f'{YC[y]}29'].number_format=CUR
r_tax=Lrow(24,"Income tax",lambda y: f"=-{YC[y]}{r_taxable}*{AB(a_tax)}")
r_npat=Lrow(25,"NPAT",lambda y: f"={YC[y]}{r_npbt}+{YC[y]}{r_tax}",bold=True)
# memo founder
L['A31']="MEMO — founder time sensitivity"; L['A31'].font=Font(bold=True, italic=True)
L['A32']="Full founder cost ($84K/yr)"
for y in YEARS: L[f'{YC[y]}32']=f"=-{AB(a_founder)}/16800*84000"; L[f'{YC[y]}32'].number_format=CUR
L['A33']="NPAT after full founder time"
for y in YEARS: L[f'{YC[y]}33']=f"={YC[y]}{r_npat}+{YC[y]}32+{AB(a_founder)}"; L[f'{YC[y]}33'].number_format=CUR
L['H32']="Founder time is a sensitivity, not a headline. Production 30 days already in the fixed block; this memo adds the other 120 days ($67,200)."
L['H32'].font=GREY
PL=dict(rev=r_prev,npat=r_npat,dep=r_dep)
def LY(r,y): return ref("P&L",f"{YC[y]}{r}")

# ============================ CASH FLOW ============================
C=wb.create_sheet("CashFlow")
C.column_dimensions['A'].width=40
for c in ['B','C','D','E','F','G']: C.column_dimensions[c].width=13
C['A1']="Cash Flow (AUD)"; C['A1'].font=HDR
for y in YEARS: C[f'{YC[y]}3']=f"Year {y}"; C[f'{YC[y]}3'].font=Font(bold=True)
def Crow(r,label,fn,fmt=CUR,bold=False,sec=False):
    C[f'A{r}']=label
    if sec: C[f'A{r}'].font=Font(bold=True, italic=True)
    if fn:
        for y in YEARS: C[f'{YC[y]}{r}']=fn(y); C[f'{YC[y]}{r}'].number_format=fmt;
        if bold:
            for y in YEARS: C[f'{YC[y]}{r}'].font=TOT
    return r
Crow(4,"OPERATING",None,sec=True)
Crow(5,"NPAT",lambda y: f"={LY(PL['npat'],y)}")
Crow(6,"add: depreciation",lambda y: f"={BY(BW['dep'],y)}")
# WC movements: -(AR_y-AR_{y-1}) etc; Y1 uses opening 0
def delta(r,y):
    if y==1: return f"{BY(r,1)}-0"
    return f"{BY(r,y)}-{BY(r,y-1)}"
Crow(7,"less: increase in receivables",lambda y: f"=-({delta(BW['ar'],y)})")
Crow(8,"less: increase in inventory",lambda y: f"=-({delta(BW['inv'],y)})")
Crow(9,"add: increase in payables",lambda y: f"=({delta(BW['ap'],y)})")
r_cfo=Crow(10,"Net cash from operating",lambda y: f"=SUM({YC[y]}5:{YC[y]}9)",bold=True)
Crow(12,"INVESTING",None,sec=True)
r_cfi=Crow(13,"Capex (container builds)",lambda y: f"=-{BY(BW['capex'],y)}",bold=True)
Crow(15,"FINANCING",None,sec=True)
Crow(16,"Debt draws",lambda y: f"={AY(a_debtdraw,y)}")
Crow(17,"Equity contributions",lambda y: f"={AY(a_equitydraw,y)}")
r_cff=Crow(18,"Net cash from financing",lambda y: f"={YC[y]}16+{YC[y]}17",bold=True)
r_net=Crow(20,"Net change in cash",lambda y: f"={YC[y]}{r_cfo}+{YC[y]}{r_cfi}+{YC[y]}{r_cff}",bold=True)
C['A21']="Cash at start"
C['C21']=f"={AB(a_opencash)}"
for y in [2,3,4,5]: C[f'{YC[y]}21']=f"={YC[y-1]}22"
for y in YEARS: C[f'{YC[y]}21'].number_format=CUR
r_cashend=22; C['A22']="Cash at end"
for y in YEARS: C[f'{YC[y]}22']=f"={YC[y]}21+{YC[y]}{r_net}"; C[f'{YC[y]}22'].number_format=CUR; C[f'{YC[y]}22'].font=TOT
def CY(r,y): return ref("CashFlow",f"{YC[y]}{r}")

# ---- Balance sheet statement (top, rows 4-26) referencing workings + CF + P&L ----
def Brow(r,label,fn,bold=False,sec=False,opening=None):
    B[f'A{r}']=label
    if sec: B[f'A{r}'].font=Font(bold=True, italic=True)
    if opening is not None: B[f'B{r}']=opening; B[f'B{r}'].number_format=CUR
    if fn:
        for y in YEARS: B[f'{YC[y]}{r}']=fn(y); B[f'{YC[y]}{r}'].number_format=CUR
        if bold:
            for y in YEARS: B[f'{YC[y]}{r}'].font=TOT
    return r
Brow(4,"ASSETS",None,sec=True)
B['B5']=f"={AB(a_opencash)}";
r_cash=Brow(5,"Cash",lambda y: f"={CY(r_cashend,y)}",opening=None); B['B5']=f"={AB(a_opencash)}"; B['B5'].number_format=CUR
r_bar=Brow(6,"Accounts receivable",lambda y: f"={BY(BW['ar'],y)}",opening=0)
r_binv=Brow(7,"Inventory",lambda y: f"={BY(BW['inv'],y)}",opening=0)
B['B8']=f"={AB(a_openplant)}"
r_bppe=Brow(8,"Net PP&E",lambda y: f"={BY(BW['netppe'],y)}"); B['B8']=f"={AB(a_openplant)}"; B['B8'].number_format=CUR
r_bassets=Brow(9,"TOTAL ASSETS",lambda y: f"=SUM({YC[y]}5:{YC[y]}8)",bold=True); B['B9']=f"=SUM(B5:B8)"; B['B9'].number_format=CUR; B['B9'].font=TOT
Brow(11,"LIABILITIES",None,sec=True)
r_bap=Brow(12,"Accounts payable",lambda y: f"={BY(BW['ap'],y)}",opening=0)
r_bdebt=Brow(13,"Borrowings",lambda y: f"={BY(BW['debt'],y)}",opening=0)
r_bliab=Brow(14,"TOTAL LIABILITIES",lambda y: f"={YC[y]}12+{YC[y]}13",bold=True); B['B14']="=B12+B13"; B['B14'].number_format=CUR; B['B14'].font=TOT
Brow(16,"EQUITY",None,sec=True)
B['B17']=f"={AB(a_opencapital)}"
r_bcontrib=Brow(17,"Contributed capital",lambda y: None);
B['B17']=f"={AB(a_opencapital)}"; B['B17'].number_format=CUR
B['C17']=f"=B17+{AY(a_equitydraw,1)}"
for y in [2,3,4,5]: B[f'{YC[y]}17']=f"={YC[y-1]}17+{AY(a_equitydraw,y)}"
for y in YEARS: B[f'{YC[y]}17'].number_format=CUR
r_bre=18; B['A18']="Retained earnings"; B['B18']=0; B['B18'].number_format=CUR
B['C18']=f"=B18+{LY(PL['npat'],1)}"
for y in [2,3,4,5]: B[f'{YC[y]}18']=f"={YC[y-1]}18+{LY(PL['npat'],y)}"
for y in YEARS: B[f'{YC[y]}18'].number_format=CUR
r_beq=Brow(19,"TOTAL EQUITY",lambda y: f"={YC[y]}17+{YC[y]}18",bold=True); B['B19']="=B17+B18"; B['B19'].number_format=CUR; B['B19'].font=TOT
B['A20']="TOTAL LIABILITIES + EQUITY";
for y in YEARS: B[f'{YC[y]}20']=f"={YC[y]}14+{YC[y]}19"; B[f'{YC[y]}20'].number_format=CUR; B[f'{YC[y]}20'].font=TOT
B['B20']="=B14+B19"; B['B20'].number_format=CUR; B['B20'].font=TOT
B['A22']="BALANCE CHECK (must be 0)"; B['A22'].font=Font(bold=True)
for y in YEARS: B[f'{YC[y]}22']=f"={YC[y]}9-{YC[y]}20"; B[f'{YC[y]}22'].number_format=CUR
B['B22']="=B9-B20"; B['B22'].number_format=CUR

# ============================ SOURCES ============================
S=wb.create_sheet("Sources")
S.column_dimensions['A'].width=40; S.column_dimensions['B'].width=90
S['A1']="Sources & status"; S['A1'].font=HDR
notes=[
 ("Full figure set","goc-data-and-figures-pack.csv (116+ rows, status + source per figure) — this workbook's inputs trace to it."),
 ("Price, BOM, path costs","Verified from invoices (Defy/DNA Steel/Centre Canvas). Path marginal costs modelled from verified inputs."),
 ("Factory path","PROVEN: 40 Maningrida Stretch beds (INV-0303) pressed + CNC'd + assembled in-house at the farm. Per-bed cost modelled pending measured actuals."),
 ("Fixed block $109,500","GoC Q&A Q1 confirms line by line."),
 ("Capex","Central factory $112-222K gross (modelled midpoints, NOT firm quotes). Sunk plant ~$75K evidenced (MVF). Container build $207,450 = Matt Inputs B27."),
 ("Community wraparound (Pot 2)","DEWR-derived: bare ~$152K/yr, staffed ~$342K/yr, program $300K/yr. Zeroed in base case; grant-funded, never bed sales."),
 ("Capital stack","PLACEHOLDER: SEFA $300K + QBE $400K (debt) + philanthropy $500K (equity). $0 signed match-eligible today. Confirm split + QBE match ratio."),
 ("Historical base","Accountant-signed Goods-only carve-out revenue $713,827 (only externally-quotable figure). ~89% grant-funded."),
 ("Caveats","Xero-mirror workpaper, NOT accountant-reviewed statutory accounts. Capex not firm-quoted. Two pots never cross-subsidise on paper."),
 ("Model conventions","Accrual P&L; capex capitalised + straight-line depreciated (full year in year of acquisition); loss carry-forward tax; debt interest-only in the 5-yr window; over-capitalised base case builds idle cash (runway)."),
]
for i,(a,b) in enumerate(notes):
    S[f'A{3+i}']=a; S[f'A{3+i}'].font=Font(bold=True)
    S[f'B{3+i}']=b; S[f'B{3+i}'].alignment=Alignment(wrap_text=True)

for sh in wb.worksheets:
    sh.sheet_view.showGridLines=True
    sh.freeze_panes="B4" if sh.title!="Sources" else "A3"

wb.save(OUT)
print("Saved", OUT)
