#!/usr/bin/env python3
"""GOC Entity Model Inputs — the simple figures sheet for Matt, mirroring the
format of his 'GOC Bed Unit-Costing Model v2.xlsx' (single Inputs page, numbered
blocks, blue = input, per-line source/status, All AUD ex-GST + a Notes sheet).
This hands Matt the ENTITY-level figures to wrap the bed model into a GOC-only
3-statement model. It does NOT build the statements — that's his build."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT="/Users/benknight/Code/Goods Asset Register/wiki/outputs/2026-07-23-goc-financial-model-pack/GOC-Entity-Model-Inputs.xlsx"
wb=openpyxl.Workbook()

BLUE=Font(color="0000CC")                 # input (matches Matt's convention)
BLACK=Font(color="000000")
BOLD=Font(bold=True)
TITLE=Font(bold=True, size=13)
SEC=Font(bold=True, color="FFFFFF", size=11)
SECFILL=PatternFill("solid", fgColor="2E5395")
GREY=Font(color="666666", size=9)
CUR='#,##0'; TWO='0.00'; PC='0%'

S=wb.active; S.title="Inputs"
S.column_dimensions['A'].width=46
S.column_dimensions['B'].width=14
S.column_dimensions['C'].width=17
S.column_dimensions['D'].width=66
r=1
S['A1']="Goods on Country — GOC Entity Model Inputs"; S['A1'].font=TITLE
S['A2']=("Blue = editable input.  Black = formula.  All AUD ex-GST unless noted.  These are the ENTITY-level figures to wrap the bed "
         "unit-costing model into a GOC-only 3-statement model. The bed model (cost/bed, facility build, production & sales, capital "
         "stack) plugs in as the production driver; this sheet adds the historical base, opening position, working capital, financing "
         "split, tax and the grant-funded community wraparound. Every figure carries a status + source. See the Notes tab for provenance.")
S['A2'].font=GREY; S['A2'].alignment=Alignment(wrap_text=True); S.merge_cells('A2:D2'); S.row_dimensions[2].height=58
r=4

def sec(title):
    global r
    S[f'A{r}']=title; S[f'A{r}'].font=SEC; S[f'A{r}'].fill=SECFILL
    for c in ['B','C','D']: S[f'{c}{r}'].fill=SECFILL
    r+=1
def hdr():
    global r
    for c,t in [('A','Line item'),('B','Value'),('C','Status'),('D','Source / notes')]:
        S[f'{c}{r}']=t; S[f'{c}{r}'].font=BOLD
    r+=1
def row(label,val,status,note,fmt=CUR,inp=True,total=False):
    global r
    S[f'A{r}']=label
    S[f'B{r}']=val
    S[f'B{r}'].number_format=fmt
    S[f'B{r}'].font=(BLUE if inp else BOLD)
    if total: S[f'A{r}'].font=BOLD; S[f'B{r}'].font=BOLD
    S[f'C{r}']=status; S[f'C{r}'].font=GREY
    S[f'D{r}']=note; S[f'D{r}'].font=GREY; S[f'D{r}'].alignment=Alignment(wrap_text=True)
    rr=r; r+=1; return rr
def frm(label,formula,note,fmt=CUR):
    global r
    S[f'A{r}']=label; S[f'A{r}'].font=BOLD
    S[f'B{r}']=formula; S[f'B{r}'].number_format=fmt; S[f'B{r}'].font=BOLD
    S[f'D{r}']=note; S[f'D{r}'].font=GREY
    rr=r; r+=1; return rr
def gap():
    global r; r+=1

# 1. HISTORICAL ACTUALS (base year)
sec("1. HISTORICAL ACTUALS — GOODS CARVE-OUT (base year, FY2026)"); hdr()
row("Revenue (accountant-signed Goods-only carve-out)",713827,"accountant-signed","The ONLY externally-quotable revenue figure. Never the $403,901 'surplus' (entity P&L is a net loss).")
r_cash=row("Cash received (ACCREC paid)",649711,"verified","ACT-infra Xero, ACT-GD, 2026-05-29.")
row("Accounts receivable (still due)",143000,"workpaper","")
r_exp=row("Goods expenses (single cash basis)",309126,"verified","ACT-GD bank spend less 1300-Washer reclass. AP genuinely ~$0 owed (matching gap, not debt).")
r_surplus=frm("Operating surplus before founder time",f"=B{r_cash}-B{r_exp}","= cash received − expenses. Model does NOT only work because founders unpaid.")
row("Grant-funded share of revenue",0.89,"verified","Grant-funded social enterprise building toward commercial sustainability.",PC)
gap()

# 2. OPENING BALANCE SHEET POSITION
sec("2. OPENING BALANCE SHEET POSITION (start of forecast)"); hdr()
r_opencash=row("Opening cash",50000,"assumption","Most important single input (base placeholder). Confirm with Ben.")
r_plant=row("Plant & equipment (sunk, evidenced)",75000,"evidenced","MVF hardware, mostly used deals (Nic). Ownership handover targeted end-Aug 2026.")
row("Borrowings (today)",0,"verified","No debt; AP ~$0 owed.")
frm("Opening contributed capital",f"=B{r_opencash}+B{r_plant}","= opening cash + sunk plant (so the opening balance sheet balances).")
gap()

# 3. GOC OPERATING COST STRUCTURE (Pot 1)
sec("3. GOC OPERATING COST STRUCTURE — per year (Pot 1, production)"); hdr()
r_fb=row("Founder production time (30d x $560)",16800,"locked","Locked 2026-05-29. Only production days touch unit cost.")
row("Kirmos facility (50% to beds)",27000,"verified","GoC Q&A Q1 confirms.")
row("Admin",14700,"verified","")
r_ft=row("Field travel",51000,"verified","")
frm("TOTAL fixed operating block / yr",f"=SUM(B{r_fb}:B{r_ft})","= $109,500, confirmed line by line (GoC Q&A Q1).")
row("Equipment maintenance (% of $167K equip, per facility)",0.05,"placeholder","3-5% pending vendor quotes. $167K = shredder+press+CNC+bench midpoints.",PC)
gap()

# 4. COMMUNITY FACILITY & PROGRAM (Pot 2)
sec("4. COMMUNITY FACILITY & PROGRAM — per site per year (Pot 2, GRANT-FUNDED, kept separate)"); hdr()
row("Bare facility operating (rent+insurance+admin+upkeep)",151666,"derived","From the Oonchiumpa DEWR application budget lines.")
row("Fully staffed facility (adds manager + trainer/WHS)",341666,"derived","Bare + manager $150K + trainer/WHS ~$40K.")
row("Employment program (brokerage/wages)",300000,"as-written","DEWR application. Grant/government money, NEVER carried by bed sales.")
S[f'A{r}']="NOTE — the two pots"; S[f'A{r}'].font=BOLD
S[f'D{r}']="Production (Pot 1) pays for itself at volume. The wraparound (Pot 2) is grant-funded by design. Keep them in separate cost centres that never cross-subsidise on paper."
S[f'D{r}'].font=GREY; S[f'D{r}'].alignment=Alignment(wrap_text=True); r+=1
gap()

# 5. CAPITAL STACK & FINANCING
sec("5. CAPITAL STACK & FINANCING"); hdr()
r_cs1=row("SEFA (debt)",300000,"placeholder","First lender target; ownership-gate-free. Draw Yr1.")
row("QBE (debt, repayable preferred)",400000,"placeholder","At-least-matched by signed external capital. Draw Yr2. Match ratio + eligible/match split TBC.")
r_cs3=row("Anchor philanthropy (equity / grant to capital)",500000,"placeholder","Draw Yr1.")
frm("TOTAL capital stack",f"=SUM(B{r_cs1}:B{r_cs3})","")
row("Interest rate on debt",0.06,"assumption","Interest-only assumed in the forecast window.",PC)
row("Signed match-eligible capital today",0,"verified","0 open Goods-eligible grants; 51% First Nations ownership = biggest unlock.")
gap()

# 6. WORKING CAPITAL & TAX
sec("6. WORKING CAPITAL & TAX (for the 3-statement build)"); hdr()
row("Debtor days",30,"assumption","On product revenue.","0")
row("Creditor days",30,"assumption","On cash operating costs.","0")
row("Inventory days",45,"assumption","On bed COGS.","0")
row("Company tax rate",0.25,"assumption","Pty Ltd base rate. Confirm charity/DGR treatment with accountant. Model loss carry-forward.",PC)
row("Plant useful life (depreciation, years)",10,"assumption","Straight-line.","0")
gap()

# 7. BED ECONOMICS (the production plug — from the bed model)
sec("7. BED ECONOMICS — the production plug (from the bed unit-costing model)"); hdr()
row("Selling price per bed",750,"verified","Shop stretch-bed-single, held flat.")
row("Marginal cost per bed — Buy-Kit",684.79,"modelled","Today's bought-in path.",TWO)
row("Marginal cost per bed — Factory",425.74,"modelled (proven)","PROCESS PROVEN: 40 Maningrida beds made in-house at the farm. Per-bed cost modelled pending measured actuals.",TWO)
row("Marginal cost per bed — Community",420.74,"modelled (future)","On-country community path not yet run.",TWO)
row("Contribution per bed — Factory / Community",329.26,"derived","$750 − Community marginal (Factory = $324.26). In-sourcing legs is the whole prize.",TWO)
row("Break-even — Factory / Community (beds/yr)",338,"derived","Fixed block ÷ contribution. Community ~333.","0")
gap()

# KEY NOTES
sec("KEY NOTES (carry into the model's cover)")
for i,t in enumerate([
 "Two pots: production (Pot 1) self-funds at volume; the community wraparound (Pot 2) is grant-funded by design, never from bed sales.",
 "Factory path is PROVEN — 40 Maningrida Stretch beds pressed + CNC'd + assembled in-house at the farm. Open = the measured per-bed cost/rate at volume, and the separate on-country community path.",
 "These are Xero-mirror workpaper figures, NOT accountant-reviewed statutory accounts. ~89% grant-funded. Capex not firm-quoted.",
 "Base year revenue to use = the accountant-signed carve-out $713,827 only.",
]):
    S[f'A{r}']=f"• {t}"; S[f'A{r}'].font=GREY; S[f'A{r}'].alignment=Alignment(wrap_text=True)
    S.merge_cells(f'A{r}:D{r}'); S.row_dimensions[r].height=28; r+=1

S.freeze_panes="A4"

# ============================ NOTES SHEET ============================
N=wb.create_sheet("Notes")
N.column_dimensions['A'].width=110
nr=1
def nput(txt, bold=False):
    global nr
    N[f'A{nr}']=txt; N[f'A{nr}'].font=(BOLD if bold else GREY)
    N[f'A{nr}'].alignment=Alignment(wrap_text=True); nr+=1
nput("Provenance, sources & status — GOC Entity Model Inputs (2026-07-23)", True)
nput("")
nput("PURPOSE", True)
nput("The entity-level figures needed to wrap the bed unit-costing model into a GOC-only 3-statement model. Mirrors the format of 'GOC Bed Unit-Costing Model v2.xlsx'. Full traceable figure set (117 rows) = goc-data-and-figures-pack.csv in the same folder.")
nput("")
nput("STATUS VOCABULARY", True)
nput("accountant-signed = endorsed by the accountant.  verified = traced to invoice / register / Xero pull / ASIC.  evidenced = Xero bill/bank-line level.  derived = arithmetic here from cited inputs.  workpaper = unaudited Xero-mirror.  modelled = calculated from verified inputs (Factory process proven, per-bed cost not yet measured at volume).  assumption / placeholder = no second source or an explicit placeholder pending a decision.")
nput("")
nput("SOURCES", True)
for s in [
 "Block 1 (historical): matt-document-bundle/04-verified-financials.md (Xero ACT-GD, 2026-05-29) + current canon revenue carve-out $713,827.",
 "Block 2 (opening position): 2026-07-22-minimal-viable-facility-model.md (sunk ~$75K evidenced).",
 "Block 3 (fixed block): 03-cost-model-and-build-paths.md; GoC Q&A Q1 confirms the $109,500 breakdown.",
 "Block 4 (Pot 2): 2026-07-22-community-facility-operating-model.md (Oonchiumpa DEWR application budget lines).",
 "Block 5 (capital stack): 04-verified-financials.md; QBE Stage 2 terms. PLACEHOLDER split.",
 "Block 6 (working capital/tax): modelling assumptions, flagged.",
 "Block 7 (bed economics): 03-cost-model-and-build-paths.md (v6). Reconciles exactly to Matt's Inputs block 1.",
]: nput("• "+s)
nput("")
nput("DECISIONS FOR BEN (the six that firm up the placeholders)", True)
for d in [
 "1. HDPE per-bed mass & rate — confirm 20kg @ $2.75 as the model figure, reconcile the '20kg diverted' claim.",
 "2. Site capex — adopt sunk ~$75K / replication ~$105K; retire the $30K and $100-150K figures.",
 "3. Capital ask — quote gross $112-222K; present sunk spend as evidence, don't net off. Never quote '$90-200K'.",
 "4. Capital stack — confirm the real split and the QBE match ratio ($0 signed today).",
 "5. Community-facility block — agree bare ~$152K / staffed ~$342K / program $300K as the three-cost-centre structure.",
 "6. Measured run — the process is proven (40 beds); confirm whether that run's actuals (time/diesel/yield) were captured, or run a short measured one to lock $425.74.",
]: nput("• "+d)
nput("")
nput("DELIBERATE EXCLUSIONS / FLAGS", True)
for f in [
 "Not accountant-reviewed statutory accounts; ~89% grant-funded.",
 "Container & equipment capex are modelled midpoints / desktop estimates, NOT firm quotes.",
 "CRM pipeline (~$3.42M) is internal only — not committed capital, not QBE-match evidence.",
 "Founder non-production time ($67,200/yr) excluded from unit cost (cost of capital, not production).",
 "Community wraparound (Pot 2) is grant-funded and separate — never funded from bed sales.",
]: nput("• "+f)

wb.save(OUT)
print("Saved", OUT)
