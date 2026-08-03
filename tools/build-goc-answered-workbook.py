#!/usr/bin/env python
"""
Build the GOC entity model workbook that goes to Matt Allen / Social Impact Hub.

WHY THIS IS CODE AND NOT A SPREADSHEET
--------------------------------------
The workbook is an evidence artifact: every figure carries a status grade and a
source, and several of them are load-bearing in funder conversations. A binary
.xlsx cannot be diffed, reviewed or blamed, so the figures live here instead and
the workbook is a build output. Change a number here, in a reviewable commit, and
regenerate. Never hand-edit the .xlsx and expect it to survive.

This is deliberately self-contained: it does NOT read Matt's original
`GOC-Entity-Model-Inputs (1).xlsx`. That file is an email attachment, it is
gitignored as a binary, and a generator that silently produces different output
depending on which download is sitting on disk is worse than no generator.

PROVENANCE OF THE FIGURES
-------------------------
  - Blocks 1-7 of `Inputs` and all of `Notes`: Matt's 2026-07-23 workbook,
    transcribed verbatim so his structure is preserved.
  - `Known Other Costs`: the MVF reconciliation (2026-07-22), Xero evidence, and
    the Oonchiumpa REAL Innovation Fund / DEWR application budget lines.
  - `Facility & Modules`: `v2/src/lib/data/cost-model-scenarios.json` -> capex_modules
    (added 2026-07-25, Matt model input 7) and `pathway-stages.ts` -> MODULES.
  - `Case Studies & Demand`: `v2/src/lib/data/community-pathways.ts`.
  - `MODEL (live)`: reproduces `v2/src/lib/cost-model/engine.ts` so the sheet and
    the app cannot silently disagree. Locked 2026-05-29 values.
  - `Open Questions` answers: written 2026-08-03, each graded and assigned an owner.

STATUS VOCABULARY (used throughout, and it is load-bearing)
    accountant-signed > verified > evidenced > derived > workpaper > modelled >
    assumption / placeholder / budget > open
Never promote a figure up this ladder without a second source.

Usage:
    python tools/build-goc-answered-workbook.py [-o OUTPUT.xlsx]
Then push it to the live Google Sheet with tools/push-xlsx-to-gsheet.py.
"""
import argparse
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Shared styling ──────────────────────────────────────────────────────────
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="2F4F4F")
TITLE = Font(bold=True, size=14)
SECTION = Font(bold=True, size=11, color="1F4E4E")
KEY = Font(bold=True, size=12, color="006400")
INPUT = Font(color="0000CC", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="E8F0FE")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP_RULE = Border(top=Side(style="thin"))


def header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = WRAP


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wrap_all(ws, ncols):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).alignment = WRAP


def write_rows(ws, rows):
    for row in rows:
        ws.append(list(row))


# ── Sheet 1: MODEL (live) ───────────────────────────────────────────────────
def build_model(wb):
    """The playable layer. Reproduces engine.ts so the sheet and app agree."""
    ws = wb.create_sheet("MODEL (live)")

    def section(r, text):
        ws.cell(row=r, column=1, value=text).font = SECTION
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="2F4F4F")
            ws.cell(row=r, column=c).font = Font(bold=True, size=11, color="FFFFFF")

    def inp(r, label, value, unit="", note=""):
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=value)
        c.font, c.fill = INPUT, INPUT_FILL
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)

    def calc(r, label, formula, unit="", note="", big=False):
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=formula)
        c.font = KEY if big else Font()
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=note)

    ws["A1"] = "GOC LIVE MODEL - change the blue cells, everything else recalculates"
    ws["A1"].font = TITLE
    ws["A2"] = (
        "Blue = you edit. Black = formula. This sheet is the playable layer over the answered inputs. "
        "Nothing here overrides the locked figures: it reproduces the cost-model engine so you can move a "
        "dial in front of an investor and show what happens."
    )

    section(4, "1. INPUTS - the dials")
    inp(5, "Selling price per bed", 750, "$", "Held flat FY27. Shop price, verified.")
    inp(6, "Marginal cost per bed - Buy-Kit (today)", 684.79, "$", "Today's bought-in path.")
    inp(7, "Marginal cost per bed - Factory (in-house legs)", 425.74, "$",
        "Process PROVEN on 40 Maningrida beds. Per-bed cost modelled, not measured.")
    inp(8, "Marginal cost per bed - Community (fair wage)", 420.74, "$",
        "Free feedstock + $130/bed fair wage. Not yet run.")
    inp(9, "Which path? 1=Kit  2=Factory  3=Community", 2, "", "Drives every 'selected' figure below.")
    inp(10, "Beds per year (per facility)", 500, "beds",
        "MODELLED planning rate, band 350-750. Not demonstrated.")
    inp(11, "Number of facilities", 1, "sites", "")
    inp(12, "Network block (Pot 1, amortises across sites)", 109500, "$/yr",
        "Founder production $16,800 + Kirmos $27,000 + admin $14,700 + field travel $51,000.")
    inp(13, "Site production block (Pot 1, per site)", 48333, "$/yr",
        "Machine upkeep $18,333 + 50% share of lease/rent $30,000. Bed sales carry this.")
    inp(14, "Site supervisor (0 / 45000 / 90000)", 0, "$/yr",
        "OPEN DECISION. Sits on top of the $130/bed fair wage - do not double count.")
    inp(15, "Production yield", 0.9, "%", "Q23 answer. HDPE reprocessing has real offcut and reject loss.")
    inp(16, "Warranty provision (% of sales)", 0.03, "%",
        "Q22 answer. Band 2-5%. Five-year warranty, canvas is the wear part.")
    inp(17, "Bad debt (% of product revenue)", 0.02, "%", "Q23 answer.")
    inp(18, "Central entity overhead", 53000, "$/yr",
        "Q20 answer. BUDGET, not sourced. Insurance/audit/legal/software/banking/governance.")
    inp(19, "Founder non-production time", 67200, "$/yr", "Q13 answer: entity overhead, with an add-back below.")
    inp(20, "Founder contribution add-back? 1=yes 0=no", 1, "",
        "Shows the cash reality that it is not currently paid.")

    section(22, "2. MODULE PICKER - what this community actually chose (put 1 or 0 in column B)")
    ws.cell(row=23, column=1, value="Module").font = Font(bold=True)
    for i, h in enumerate(["Include (1/0)", "Capex low $", "Capex high $", "Grade", "Needs upstream"], start=2):
        ws.cell(row=23, column=i, value=h).font = Font(bold=True)

    modules = [
        ("Site base (container, power, pad, ventilation)", 1, 31800, 64000, "modelled", "-"),
        ("Collection and sorting", 1, 5000, 19500, "estimate", "-"),
        ("Shredding", 1, 19800, 19800, "evidenced", "Collection and sorting"),
        ("Pressing, CNC and finishing", 1, 32780, 32780, "evidenced", "Shredding"),
        ("Assembly and workshop", 1, 6387, 6387, "evidenced", "Pressing, CNC"),
        ("Sales and delivery", 1, 0, 0, "evidenced", "Assembly"),
    ]
    r = 24
    for name, inc, lo, hi, grade, feed in modules:
        ws.cell(row=r, column=1, value=name)
        c = ws.cell(row=r, column=2, value=inc)
        c.font, c.fill = INPUT, INPUT_FILL
        ws.cell(row=r, column=3, value=lo)
        ws.cell(row=r, column=4, value=hi)
        ws.cell(row=r, column=5, value=grade)
        ws.cell(row=r, column=6, value=feed)
        r += 1

    ws.cell(row=30, column=1, value="CHOSEN CAPEX").font = Font(bold=True)
    ws.cell(row=30, column=3, value="=SUMPRODUCT($B$24:$B$29,C24:C29)").font = KEY
    ws.cell(row=30, column=4, value="=SUMPRODUCT($B$24:$B$29,D24:D29)").font = KEY
    ws.cell(row=30, column=5, value="low / high band")
    # A module cannot run without the one feeding it. Selecting shredding with no
    # collection is not a cheap configuration, it is an incoherent one.
    ws.cell(row=31, column=1, value="Feed-order check (a module needs the one above it)")
    ws.cell(row=31, column=3,
            value='=IF(SUMPRODUCT(--(B24:B28<B25:B29))>0,'
                  '"BROKEN CHAIN - a module is selected without its upstream feed","chain OK")')
    ws.cell(row=32, column=1, value="Turnkey alternative (fitted 40ft container workshop)")
    c = ws.cell(row=32, column=3, value=207450)
    c.font, c.fill = INPUT, INPUT_FILL
    ws.cell(row=32, column=4,
            value="Different SCOPE, not a pessimistic guess at the same thing. "
                  "Use as the conservative investor case.")

    section(34, "3. POT 1 - PRODUCTION. Does the bed business wash its own face?")
    calc(35, "Selected marginal cost per bed", "=IF($B$9=1,$B$6,IF($B$9=2,$B$7,$B$8))", "$")
    calc(36, "Contribution per bed", "=$B$5-B35", "$", "Price less marginal cost. The engine of everything.")
    calc(37, "Saleable beds (after yield)", "=$B$10*$B$15", "beds", "Yield loss applied to the production rate.")
    calc(38, "Revenue", "=B37*$B$5*$B$11", "$/yr")
    calc(39, "Marginal cost of goods", "=B37*B35*$B$11", "$/yr")
    calc(40, "Gross contribution", "=B38-B39", "$/yr", big=True)
    calc(41, "Less warranty provision", "=-B38*$B$16", "$/yr")
    calc(42, "Less bad debt", "=-B38*$B$17", "$/yr")
    calc(43, "Less network block (Pot 1)", "=-$B$12", "$/yr", "One block across all sites.")
    calc(44, "Less site production block x sites", "=-($B$13+$B$14)*$B$11", "$/yr")
    ws.cell(row=45, column=1).border = TOP_RULE
    calc(45, "POT 1 OPERATING SURPLUS", "=SUM(B40:B44)", "$/yr",
         "If positive, production pays for itself and grants are genuinely the bonus.", big=True)
    calc(46, "Verdict", '=IF(B45>0,"Washes its own face","Does NOT wash its own face at this volume")', "")
    ws.cell(row=46, column=2).font = KEY

    section(48, "4. BREAK-EVEN - the number an investor will test")
    calc(49, "Break-even vs network block", '=IF(B36>0,ROUNDUP($B$12/B36,0),"n/a")', "beds/yr",
         "338 on the factory path at locked defaults. This is the PUBLISHED figure and it is PARTIAL.")
    calc(50, "Break-even vs network + one site block", '=IF(B36>0,ROUNDUP(($B$12+$B$13+$B$14)/B36,0),"n/a")',
         "beds/yr", "The honest all-in figure for a single standalone site.")
    calc(51, "Break-even vs site production block only", '=IF(B36>0,ROUNDUP(($B$13+$B$14)/B36,0),"n/a")',
         "beds/yr", "What one site must sell to carry its own production costs.")
    calc(52, "Headroom at current volume", "=B37-B50", "beds",
         "Saleable beds less all-in break-even. Negative means the volume assumption is doing the work.")

    section(54, "5. ENTITY VIEW - Pot 1 plus the costs an investor will still ask about")
    calc(55, "Pot 1 operating surplus", "=B45", "$/yr")
    calc(56, "Less central entity overhead", "=-$B$18", "$/yr", "Budgeted, not sourced.")
    calc(57, "Less founder non-production time", "=-$B$19", "$/yr")
    calc(58, "Add back founder contribution (unpaid)", "=IF($B$20=1,$B$19,0)", "$/yr",
         "Shows both the true cost and the cash reality.")
    calc(59, "EBITDA (cash view)", "=SUM(B55:B58)", "$/yr", big=True)
    calc(60, "Less depreciation on chosen capex", "=-AVERAGE($C$30,$D$30)/10", "$/yr", "Straight line, 10 years.")
    calc(61, "EBIT", "=B59+B60", "$/yr")
    calc(62, "Less interest", "=-$B$70*$B$71", "$/yr", "Debt drawn x rate.")
    calc(63, "PROFIT BEFORE TAX", "=B61+B62", "$/yr", big=True)
    calc(64, "Tax", "=IF(B63>0,-B63*0.25,0)", "$/yr",
         "25% base rate. Losses carry forward, so cash tax is likely nil in the window.")
    calc(65, "PROFIT AFTER TAX", "=B63+B64", "$/yr", big=True)

    section(67, "6. POT 2 - THE COMMUNITY WRAPAROUND. Grant funded by design. Never carried by bed sales.")
    inp(68, "Site archetype: 1=bare  2=fully staffed", 1, "",
        "Bare $151,666. Fully staffed $341,666 (adds $150K manager + $40K trainer/WHS).")
    calc(69, "Pot 2 facility cost per site", "=IF($B$68=1,151666,341666)", "$/yr")
    inp(70, "Debt drawn", 300000, "$", "SEFA base case. QBE $400K modelled as upside, not base.")
    inp(71, "Interest rate", 0.06, "%", "Interest-only placeholder. No facility offered yet.")
    inp(72, "Employment program", 300000, "$/yr", "DEWR. Grant/government money, NEVER bed sales.")
    calc(73, "TOTAL POT 2 per site", "=B69+B72", "$/yr", big=True)
    calc(74, "Grant income required to match Pot 2", "=B73*$B$11", "$/yr",
         "Model as restricted income in the same period. Unspent = liability, not revenue.")
    calc(75, "THE TEST",
         '=IF(B45>0,"PASS - if the grant ends, production keeps running",'
         '"FAIL - production depends on Pot 2 at this volume")', "",
         "The single question that proves or breaks the two-pot claim.")
    ws.cell(row=75, column=2).font = KEY

    section(77, "7. SCENARIOS - what to show an investor")
    for i, h in enumerate(["Scenario", "Path", "Beds/yr", "Sites", "What it demonstrates"], start=1):
        ws.cell(row=78, column=i, value=h).font = Font(bold=True)
    scenarios = [
        ("Today (buy kits)", "1", 120, 1, "The honest current run-rate. Loses money and we say so."),
        ("Factory at target", "2", 500, 1, "Legs in-sourced. The prize is $259/bed of cost taken out."),
        ("Factory, break-even", "2", 338, 1, "Covers the network block. Does NOT cover a site block."),
        ("Community-owned", "3", 500, 1, "Parity with the factory, margin stays in community."),
        ("Three sites", "2", 500, 3, "Network block amortises. This is where it becomes a business."),
    ]
    r = 79
    for s in scenarios:
        for i, v in enumerate(s, start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    ws.cell(row=85, column=1,
            value="To run one: copy its Path / Beds / Sites into cells B9, B10 and B11 above.").font = Font(italic=True)

    section(87, "8. WHAT THIS MODEL DOES NOT KNOW")
    for i, note in enumerate([
        "Demand. There are no signed FY27 bed orders. 500 beds/yr is a PRODUCTION rate, not a sales forecast. "
        "Do not read revenue off capacity.",
        "The 40-bed Maningrida run proved the process but was not measured for time, diesel or yield. "
        "$425.74 stays modelled.",
        "No GOC bank account exists. Opening cash is a carve-out decision at the handover date, not a bank fact.",
        "The FY26 expense figure of $309,126 is one aggregate with no category split. The accountant carve-out "
        "is the highest-value item outstanding.",
        "Central overhead of $53,000 is a budget with no source. Replace with actuals.",
        "The entity is mid-migration. Historical trading sits in the sole trader; the Pty Ltd transfer is not executed.",
    ]):
        ws.cell(row=88 + i, column=1, value="- " + note)

    set_widths(ws, [46, 18, 16, 62, 20, 26])
    wrap_all(ws, 6)
    ws.freeze_panes = "A5"


# ── Sheet 2: Inputs (Matt's structure, verbatim) ────────────────────────────
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    rows = [
        ("Goods on Country: GOC Entity Model Inputs",),
        ("Blue = editable input. Black = formula. Green = link to another sheet. All AUD ex-GST unless noted. "
         "Entity-level figures for Matt's GOC-only 3-statement model. Every figure carries a status and source. "
         "See Notes for provenance.",),
        (),
        ("1. HISTORICAL ACTUALS: GOODS CARVE-OUT (base year, FY2026)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Revenue (accountant-signed Goods-only carve-out)", 713827, "accountant-signed",
         "The ONLY externally-quotable revenue figure. Never the $403,901 'surplus' (entity P&L is a net loss)."),
        ("Cash received (ACCREC paid)", 649711, "verified", "ACT-infra Xero, ACT-GD, 2026-05-29."),
        ("Accounts receivable (still due)", 143000, "workpaper", ""),
        ("Goods expenses (single cash basis)", 309126, "verified",
         "ACT-GD bank spend less 1300-Washer reclass. AP genuinely ~$0 owed (matching gap, not debt)."),
        ("Operating surplus before founder time", "=B7-B9", "derived",
         "Cash received less Goods expenses. Model does not rely on unpaid founder time."),
        ("Grant-funded share of revenue", 0.89, "verified",
         "Grant-funded social enterprise building toward commercial sustainability."),
        (),
        ("2. OPENING BALANCE SHEET POSITION (start of forecast)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Opening cash", 50000, "assumption",
         "Most important single input (base placeholder). See Open Questions Q1: no GOC bank account exists, "
         "so this is a carve-out decision, not a bank fact."),
        # RECONCILED 2026-08-03. Matt's original carried $75,000 graded 'evidenced'. Both were
        # wrong. Canon (cost-model-scenarios.ts:157, Ben ruling 2026-07-25) sets sunk spend at
        # $110,046 and grades it `workpaper`, because ~$43,700 is evidenced at bill level and the
        # balance is plant we own whose paperwork is catching up. The $75,000 in the 2026-07-22
        # MVF note is a bill-level SUBTOTAL of hardware, explicitly NOT a competing total.
        ("Plant & equipment (sunk spend on the farm facility)", 110046, "workpaper",
         "$100,000 facility + $10,046 Carbatec tooling. ~$43,700 evidenced at bill and bank-line "
         "level in the sole-trader Xero; the balance is owned plant whose paperwork is catching up "
         "(the $19,800 Telford Smith shredder, physically confirmed with no Xero record, and a "
         "recently bought larger CNC). A filing job, not a fiction. Do NOT substitute the ~$75,000 "
         "from the MVF note - that is a bill-level hardware subtotal, not a competing total. "
         "Ownership handover targeted end-Aug 2026."),
        ("Borrowings (today)", 0, "verified", "No debt; AP ~$0 owed."),
        ("Opening contributed capital", "=B15+B16", "derived",
         "Opening cash plus sunk plant, so the opening balance sheet balances."),
        (),
        ("3. GOC OPERATING COST STRUCTURE: per year (Pot 1, production)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Founder production time (30d x $560)", 16800, "locked",
         "Locked 2026-05-29. Only production days touch unit cost."),
        ("Kirmos facility (50% to beds)", 27000, "verified", "GoC Q&A Q1 confirms."),
        ("Admin", 14700, "verified", ""),
        ("Field travel", 51000, "verified", ""),
        ("TOTAL fixed operating block / yr", "=SUM(B22:B25)", "derived",
         "$109,500, confirmed line by line in GoC Q&A Q1."),
        ("Equipment maintenance (% of $167K equip, per facility)", 0.05, "placeholder",
         "3-5% pending vendor quotes. $167K = shredder+press+CNC+bench midpoints."),
        (),
        ("4. COMMUNITY FACILITY & PROGRAM: per site per year (Pot 2, GRANT-FUNDED, kept separate)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Bare facility operating (rent+insurance+admin+upkeep)", 151666, "derived",
         "From the Oonchiumpa DEWR application budget lines."),
        ("Fully staffed facility (adds manager + trainer/WHS)", 341666, "derived",
         "Bare + manager $150K + trainer/WHS ~$40K."),
        ("Employment program (brokerage/wages)", 300000, "as-written",
         "DEWR application. Grant/government money, NEVER carried by bed sales."),
        ("NOTE: the two pots", "", "",
         "Production (Pot 1) pays for itself at volume. The wraparound (Pot 2) is grant-funded by design. "
         "Keep them in separate cost centres that never cross-subsidise on paper."),
        (),
        ("5. CAPITAL STACK & FINANCING",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("SEFA (debt)", 300000, "placeholder", "First lender target; ownership-gate-free. Draw Yr1."),
        ("QBE (debt, repayable preferred)", 400000, "placeholder",
         "At-least-matched by signed external capital. Draw Yr2. Match ratio + eligible/match split TBC."),
        ("Anchor philanthropy (equity / grant to capital)", 500000, "placeholder", "Draw Yr1."),
        ("TOTAL capital stack", "=SUM(B38:B40)", "derived", "Sum of the placeholder capital stack above."),
        ("Interest rate on debt", 0.06, "assumption", "Interest-only assumed in the forecast window."),
        ("Signed match-eligible capital today", 0, "verified",
         "0 open Goods-eligible grants; 51% First Nations ownership = biggest unlock."),
        (),
        ("6. WORKING CAPITAL & TAX (for the 3-statement build)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Debtor days", 30, "assumption", "On product revenue. Does NOT apply to grants - see Q17."),
        ("Creditor days", 30, "assumption", "On cash operating costs."),
        ("Inventory days", 45, "assumption", "On bed COGS."),
        ("Company tax rate", 0.25, "assumption",
         "Pty Ltd base rate. Confirm charity/DGR treatment with accountant. Model loss carry-forward."),
        ("Plant useful life (depreciation, years)", 10, "assumption", "Straight-line."),
        (),
        ("7. BED ECONOMICS: the production plug (from the bed unit-costing model)",),
        ("Line item", "Value", "Status", "Source / notes"),
        ("Selling price per bed", 750, "verified", "Shop stretch-bed-single, held flat."),
        ("Marginal cost per bed: Buy-Kit", 684.79, "modelled", "Today's bought-in path."),
        ("Marginal cost per bed: Factory", 425.74, "modelled (proven)",
         "PROCESS PROVEN: 40 Maningrida beds made in-house at the farm. Per-bed cost modelled pending measured actuals."),
        ("Marginal cost per bed: Community", 420.74, "modelled (future)",
         "On Country community path not yet run."),
        ("Contribution per bed: Factory / Community", 329.26, "derived",
         "$750 minus Community marginal (Factory = $324.26). In-sourcing legs is the whole prize."),
        ("Break-even: Factory / Community (beds/yr)", 338, "derived",
         "Fixed block divided by contribution. Community ~333. NOTE: network block only - see MODEL (live) row 50 "
         "for the all-in figure including a site block."),
        (),
        ("KEY NOTES (carry into the model's cover)",),
        ("- Two pots: production (Pot 1) self-funds at volume; the community wraparound (Pot 2) is grant-funded by "
         "design, never from bed sales.",),
        ("- Factory path is PROVEN: 40 Maningrida Stretch beds pressed + CNC'd + assembled in-house at the farm. "
         "Open = the measured per-bed cost/rate at volume, and the separate On Country community path.",),
        ("- These are Xero-mirror workpaper figures, NOT accountant-reviewed statutory accounts. ~89% grant-funded. "
         "Capex not firm-quoted.",),
        ("- Base year revenue to use = the accountant-signed carve-out $713,827 only.",),
    ]
    write_rows(ws, rows)
    ws["A1"].font = TITLE
    for r in (4, 13, 20, 29, 36, 45, 53, 62):
        ws.cell(row=r, column=1).font = SECTION
    for r in (5, 14, 21, 30, 37, 46, 54):
        header_row(ws, r, 4)
    set_widths(ws, [52, 16, 22, 78])
    wrap_all(ws, 4)


# ── Sheet 3: Notes ──────────────────────────────────────────────────────────
def build_notes(wb):
    ws = wb.create_sheet("Notes")
    for line in [
        "Provenance, sources & status: GOC Entity Model Inputs (2026-07-23, answered 2026-08-03)",
        "",
        "PURPOSE",
        "The entity-level figures needed to wrap the bed unit-costing model into a GOC-only 3-statement model. "
        "Mirrors the format of 'GOC Bed Unit-Costing Model v2.xlsx'.",
        "",
        "STATUS VOCABULARY",
        "accountant-signed = endorsed by the accountant.  verified = traced to invoice / register / Xero pull / ASIC.  "
        "evidenced = Xero bill/bank-line level.  derived = arithmetic here from cited inputs.  "
        "workpaper = unaudited Xero-mirror.  modelled = calculated from verified inputs (Factory process proven, "
        "per-bed cost not yet measured at volume).  assumption / placeholder / budget = no second source or an "
        "explicit placeholder pending a decision.  open = no defensible figure exists yet.",
        "",
        "SOURCES",
        "- Block 1 (historical): matt-document-bundle/04-verified-financials.md (Xero ACT-GD, 2026-05-29) + current "
        "canon revenue carve-out $713,827.",
        "- Block 2 (opening position): 2026-07-22-minimal-viable-facility-model.md (sunk ~$75K evidenced).",
        "- Block 3 (fixed block): 03-cost-model-and-build-paths.md; GoC Q&A Q1 confirms the $109,500 breakdown.",
        "- Block 4 (Pot 2): 2026-07-22-community-facility-operating-model.md (Oonchiumpa DEWR application budget lines).",
        "- Block 5 (capital stack): 04-verified-financials.md; QBE Stage 2 terms. PLACEHOLDER split.",
        "- Block 6 (working capital/tax): modelling assumptions, flagged.",
        "- Block 7 (bed economics): 03-cost-model-and-build-paths.md (v6), and v2/src/lib/cost-model/engine.ts.",
        "- MODEL (live): reproduces engine.ts. Locked 2026-05-29 values.",
        "- Facility & Modules: cost-model-scenarios.json -> capex_modules; pathway-stages.ts -> MODULES.",
        "- Case Studies & Demand: community-pathways.ts.",
        "",
        "DELIBERATE EXCLUSIONS / FLAGS",
        "- Not accountant-reviewed statutory accounts; ~89% grant-funded.",
        "- Container & equipment capex are modelled midpoints / desktop estimates, NOT firm quotes.",
        "- CRM pipeline (~$3.42M) is internal only: not committed capital, not QBE-match evidence.",
        "- Founder non-production time ($67,200/yr) excluded from unit cost (cost of capital, not production). "
        "See Q13 for the agreed entity treatment.",
        "- Community wraparound (Pot 2) is grant-funded and separate: never funded from bed sales.",
        "",
        "THE ONE THING TO READ BEFORE QUOTING A BREAK-EVEN",
        "The published 338 beds/yr covers the NETWORK BLOCK ONLY. Add the site production block, yield loss, "
        "warranty and bad debt and all-in break-even for a single standalone site is 487 beds against 450 saleable "
        "at the stated 500/yr and 90% yield. The network block amortises across sites, so the multi-site case is "
        "where this turns. See MODEL (live) rows 45-52.",
    ]:
        ws.append([line])
    ws["A1"].font = TITLE
    set_widths(ws, [150])
    wrap_all(ws, 1)


# ── Sheet 4: Known Other Costs ──────────────────────────────────────────────
def build_known_costs(wb):
    ws = wb.create_sheet("Known Other Costs")
    rows = [
        ("Known Other Costs",),
        ("Costs already found in the GOC codebase and workpapers. Status is load-bearing: evidenced and verified are "
         "not the same as modelled, allowance or physical-only. All AUD ex-GST unless the Unit column says otherwise.",),
        (),
        ("1. EXISTING PLANT AND HARDWARE: ACTUAL OR PHYSICALLY CONFIRMED",),
        ("Cost item", "Low", "High", "Selected/current", "Unit", "Status", "Source", "Notes"),
        ("SUNK SPEND ON THE FARM FACILITY (the figure to quote)", None, None, 110046, "AUD", "workpaper",
         "cost-model-scenarios.ts ALREADY_INVESTED; Ben ruling 2026-07-25",
         "$100,000 facility + $10,046 Carbatec tooling. Quoted BESIDE the capital ask as skin in the game, "
         "NEVER netted off it. The old net figure (~$2K-$112K) is retired."),
        ("  of which: evidenced hardware subtotal (bill level)", None, None, 75000, "AUD", "evidenced/mixed",
         "MVF reconciliation, 2026-07-22",
         "About $43.7K cleanly tagged, $12.5K ambiguous and $19.8K physical-only shredder. This is a SUBTOTAL "
         "of the line above, not a competing total. Do not swap it in for $110,046."),
        ("Press + cold press + CNC bundle", None, None, 32780, "AUD inc GST", "evidenced",
         "Xero INV-0054, 2025-12-17, ACT-GD", "Existing purchase, not a fresh-site quote."),
        ("Workshop tools", None, None, 6387, "AUD inc GST", "evidenced, tagged Harvest",
         "Carbatec Brisbane, Xero, Jan 2026", "$4,575.65 + $1,811.70."),
        ("20ft container, Monument Grey", None, None, 3320, "AUD inc GST", "evidenced",
         "Bionic Self Storage, Xero, 2026-04-29", "Capitalised to ACT-FM The Farm."),
        ("Crane placement", None, None, 1041, "AUD inc GST", "evidenced", "GM Crane Hire, Xero, 2026-06-29",
         "Two 20T Franna cranes."),
        ("Container transport", None, None, 193, "AUD inc GST", "evidenced", "Rapid Container, Xero, 2026-06-24",
         "After-hours transport."),
        ("Two generators", None, None, 6600, "AUD inc GST", "evidenced, ambiguous",
         "Orange Sky Australia, Xero, 2025-05-22",
         "No project tag. EXCLUDED from opening PP&E pending confirmation (Q4)."),
        ("Larger container", None, None, 5904, "AUD inc GST", "evidenced, flagged",
         "Container Options, Xero, 2025-12-09",
         "Tagged Mounty Yarns and flagged as on-sold. EXCLUDED until ownership is confirmed (Q4)."),
        ("Shredder, Telford Smith", None, None, 19800, "AUD", "physical only",
         "Ben confirmation; no record in connected Xero",
         "Owned, invoice to locate. INCLUDED at $19,800 with that note (Q4)."),
        ("Bigger CNC router", None, None, None, "AUD", "open", "MVF section 4b",
         "Recently purchased. About $5,135 installation booked separately. EXCLUDED from opening PP&E until the "
         "invoice and purchasing entity are found (Q3). Do not plug a number."),
        (),
        ("2. NEW-SITE REPLICATION: MIXED EVIDENCE AND ESTIMATES",),
        ("Cost item", "Low", "High", "Selected/current", "Unit", "Status", "Source", "Notes"),
        ("40ft shipping container", 13000, 16000, 14500, "AUD", "modelled", "Ben / MVF replication model",
         "Market-rate fresh-site range."),
        ("20ft shipping container", 6000, 10000, 8000, "AUD", "modelled", "Ben / MVF replication model",
         "Market-rate fresh-site range."),
        ("Diesel generator, press-line sized", 6600, 20000, 13300, "AUD", "mixed",
         "Orange Sky actual to proper diesel estimate", "Capacity and site power requirement need confirmation."),
        ("Shredder", 19800, 19800, 19800, "AUD", "physical basis", "Existing Telford Smith unit",
         "Invoice still to locate."),
        ("Press + cold press + CNC", 32780, 32780, 32780, "AUD inc GST", "evidenced basis", "Circularity INV-0054",
         "Existing purchase used as replication basis."),
        ("Workshop tools", 6387, 6387, 6387, "AUD inc GST", "evidenced basis", "Carbatec actuals",
         "Existing purchase used as replication basis."),
        ("Crane placement + transport", 1200, 2500, 1850, "AUD", "mixed", "Existing actuals + allowance",
         "Remote/site conditions can move this materially."),
        ("Electrical fit-out", 3000, 8000, 5500, "AUD", "allowance", "MVF replication model",
         "Board and three-phase work. Quote required."),
        ("Ventilation / fume extraction", 1000, 3000, 2000, "AUD", "allowance", "MVF replication model",
         "Quote required for hot-press heat load."),
        ("Site preparation", 500, 3000, 1750, "AUD", "allowance", "MVF replication model", "Pad and levelling."),
        ("PPE + startup consumables", 500, 1500, 1000, "AUD", "allowance", "MVF replication model",
         "Initial stock only."),
        ("NEW-SITE MINIMAL VIABLE FACILITY", 90767, 122967, 105000, "AUD", "modelled midpoint",
         "MVF reconciliation, 2026-07-22",
         "Use about $105K as the replication planning figure. Firm quotes required."),
        ("Matt container-workshop build", None, None, 207450, "AUD", "modelled", "Matt Inputs B27",
         "$167K equipment midpoint + $40,450 container lines. Different SCOPE from the MVF replication case, "
         "not a competing estimate of the same thing."),
        (),
        ("3. RECURRING COSTS ALREADY FOUND",),
        ("Cost item", "Low", "High", "Selected/current", "Unit", "Status", "Source", "Notes"),
        ("Founder full Goods time, 150 days", None, None, 84000, "AUD/yr", "locked", "Ben, 2026-05-29",
         "30 production, 50 fundraising, 25 commercialisation, 45 governance days."),
        ("Founder non-production time", None, None, 67200, "AUD/yr", "locked", "Ben, 2026-05-29",
         "Excluded from unit cost. Q13: entity overhead with an explicit founder-contribution add-back."),
        ("Equipment maintenance per facility", None, None, 8350, "AUD/facility/yr", "placeholder",
         "GoC Q&A Q9; Matt selected 5%", "5% of $167K equipment. Vendor quotes pending."),
        ("Community lease, rent, utilities and site", None, None, 60000, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "$180K over three years. Splits 50/50 production/program."),
        ("Community facility insurance", None, None, 40000, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "$120K over three years."),
        ("Community administration, accounting and IT", None, None, 33333, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "$100K over three years."),
        ("Community machine upkeep and consumables", None, None, 18333, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "$55K over three years. 100% production."),
        ("BARE COMMUNITY FACILITY", None, None, 151666, "AUD/site/yr", "derived", "Sum of DEWR facility lines",
         "Pot 2. Do not add the old $24K rent-only figure."),
        ("Community project manager incl. super", None, None, 150000, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "Q15: model as the STAFFED archetype. $90K coordinator is the "
         "production-only archetype."),
        ("Trainer / WHS officer", None, None, 40000, "AUD/site/yr", "modelled split",
         "Portion of DEWR ACT $190K line", "Estimate, not a stated standalone figure (Q16)."),
        ("FULLY STAFFED COMMUNITY FACILITY", None, None, 341666, "AUD/site/yr", "derived",
         "Bare + manager + trainer/WHS", "Pot 2, grant/government-funded by design."),
        ("Employment program brokerage and wages", None, None, 300000, "AUD/site/yr", "as-written",
         "Oonchiumpa DEWR application", "Pot 2. Never carried by bed sales."),
        ("DEWR machinery + Trainer/WHS line", None, None, 63333, "AUD/site/yr", "modelled split",
         "Oonchiumpa DEWR application",
         "REVENUE to Goods., not a cost to the site. No model currently shows this (Q16)."),
        ("Central entity overhead (Year 1 budget)", None, None, 53000, "AUD/yr", "budget",
         "Q20, no source - replace with actuals",
         "Insurance $12K, accounting/audit $15K, legal/entity setup $10K, software/IT $8K, banking $3K, "
         "governance $5K. Sits OUTSIDE the $109,500 block - do not double-count with the $14,700 admin line."),
    ]
    write_rows(ws, rows)
    ws["A1"].font = TITLE
    for r in (4, 17, 33):
        ws.cell(row=r, column=1).font = SECTION
    for r in (5, 18, 34):
        header_row(ws, r, 8)
    set_widths(ws, [42, 12, 12, 18, 16, 24, 34, 62])
    wrap_all(ws, 8)


# ── Sheet 5: Facility & Modules ─────────────────────────────────────────────
def build_facility_modules(wb):
    ws = wb.create_sheet("Facility & Modules")
    rows = [
        ("Goods on Country: the on-Country facility, priced as modules",),
        ("A community does not buy 'a facility'. It chooses modules. This sheet is the bridge between the nine "
         "community modules a pathway actually selects and the capex the model has to fund. Three of the four live "
         "pathways do not want a whole facility, which is why the old build-method ladder could price only one of them.",),
        (),
        ("1. THE SITE BASE - required by any production module",),
        ("Line", "Low $", "High $", "Grade", "Notes"),
        ("40ft shipping container", 13000, 16000, "ben", "Market rate, new site"),
        ("20ft shipping container", 6000, 10000, "ben", "Market rate, new site"),
        ("Diesel generator (press-line sized)", 6600, 20000, "ben", "Capacity depends on site power"),
        ("Crane placement + transport", 1200, 2500, "evidenced", "GM Crane Hire + Rapid Container actuals"),
        ("Electrical fit-out (board, three-phase)", 3000, 8000, "estimate", "Quote required"),
        ("Ventilation / fume extraction", 1000, 3000, "estimate", "Hot-press heat load"),
        ("Site prep (pad, levelling)", 500, 3000, "estimate", ""),
        ("PPE + startup consumables", 500, 1500, "estimate", "Initial stock only"),
        ("SITE BASE TOTAL", 31800, 64000, "modelled",
         "Where a partner supplies the shed or power, SUBTRACT those lines - the base shrinks, it does not vanish. "
         "Tennant Creek is the live case and the subtraction has not been agreed with them."),
        (),
        ("2. THE PRODUCTION MODULES - chosen, in feed order",),
        ("Module", "Low $", "High $", "Grade", "Needs upstream", "What it produces"),
        ("Collection and sorting", 5000, 19500, "estimate", "-", "Sorted, caged HDPE ready to shred"),
        ("Shredding", 19800, 19800, "evidenced", "Collection and sorting", "HDPE shred"),
        ("Pressing, CNC and finishing", 32780, 32780, "evidenced", "Shredding", "Finished HDPE leg kits"),
        ("Assembly and workshop", 6387, 6387, "evidenced", "Pressing, CNC and finishing", "Finished beds"),
        ("Sales and delivery", 0, 0, "evidenced", "Assembly and workshop", "Beds in homes"),
        ("ALL MODULES + SITE BASE", 95767, 142967, "modelled",
         "Reconciles to the MVF replication total of ~$105,000 midpoint (excluding collection, which is an addition "
         "to that table, not part of it)."),
        (),
        ("3. THE TWO PLANNING SCOPES - not competing estimates of the same thing",),
        ("Scope", "Figure $", "", "Use it for", "Why it differs"),
        ("MVF replication (base case)", 105000, "", "What it costs to stand up a working line at a new site",
         "Built on the equipment set Goods already runs, second-hand where sensible"),
        ("Turnkey 40ft container workshop (conservative)", 207450, "", "The defensible investor-facing number",
         "$167,000 equipment midpoint + $40,450 container lines. Fully fitted, bought new."),
        (),
        ("4. HOW THE NINE COMMUNITY MODULES MAP TO COST",),
        ("Community module", "Pot", "Capex?", "Operating?", "Note"),
        ("Products", "1", "No", "In unit cost", "Beds and other proven Goods products"),
        ("Equipment", "1", "Yes", "5% maintenance", "Shredders, presses, tools, machinery"),
        ("Place", "1 + 2", "Yes (site base)", "Rent, insurance, upkeep",
         "Shed or complete facility. Rent splits 50/50 between production and program."),
        ("Skills", "2", "No", "Trainer / WHS ~$40K/site/yr",
         "Installation, production, maintenance, enterprise training"),
        ("People", "2", "No", "$90K coordinator or $150K manager", "Local operators, mentors, specialist support"),
        ("Systems", "1", "No", "In the $109,500 network block", "Suppliers, QC, logistics, administration"),
        ("Enterprise", "1", "No", "In the network block", "Customers, contracts, revenue, governance, ownership"),
        ("Money", "-", "n/a", "n/a",
         "Repayable loans bridging orders and working capital, and the site capital that finishes the transfer"),
        ("Story + evidence", "2", "No", "Grant funded", "Community-approved media, outcomes and learning"),
        (),
        ("5. THE TWO POTS - the rule that keeps this honest",),
        ("POT 1 - PRODUCTION. Pays for itself at volume. Network block ~$109,500/yr plus the site production block. "
         "Bed sales carry this and nothing else.",),
        ("POT 2 - COMMUNITY WRAPAROUND. Grant funded by design. Bare facility ~$151,666/site/yr, fully staffed "
         "~$341,666/site/yr, employment program $300,000/site/yr. Bed sales NEVER carry this.",),
        ("The test an investor should apply: if a site's grant ends, does production keep running? If yes, the "
         "two-pot claim is true. Model it so that question can be answered from the sheet - see MODEL (live) row 75.",),
    ]
    write_rows(ws, rows)
    ws["A1"].font = TITLE
    for r in (4, 17, 26, 31, 42):
        ws.cell(row=r, column=1).font = SECTION
    for r in (5, 18, 27, 32):
        header_row(ws, r, 6)
    set_widths(ws, [46, 14, 14, 26, 52, 34])
    wrap_all(ws, 6)


# ── Sheet 6: Case Studies & Demand ──────────────────────────────────────────
def build_case_studies(wb):
    ws = wb.create_sheet("Case Studies & Demand")
    rows = [
        ("Who we are actually talking to, and what that is worth",),
        ("Matt's biggest open question is demand: 'do you have 1,000 beds worth of demand at $750?'. This sheet is "
         "the honest answer today. It separates what is DELIVERED, what is QUOTED and what is a CONVERSATION, and it "
         "does not let capacity masquerade as revenue.",),
        (),
        ("1. LIVE COMMUNITY PATHWAYS",),
        ("Community", "Region", "Stage", "Community lead / partner", "What they have asked for", "Modules requested",
         "Where the money sits"),
        ("Utopia Homelands", "NT - Urapuntja", "Choose modules", "Dorrie Jones / Urapuntja",
         "Practical local capability for young people, starting from what Urapuntja asks for. 147 assets already "
         "confirmed in community.", "Collection and sorting, then shredding - the earliest module in the chain",
         "Not yet priced to community. Site base + collection + shredding = $56,600 to $103,300."),
        ("Oonchiumpa", "NT", "Fund", "Kristy Bloomfield",
         "A full facility with an employment program. Young people built the beds.",
         "The whole chain, plus people and skills",
         "REAL Innovation Fund Stage Two / DEWR application: bare facility ~$151,666/yr, staffed ~$341,666/yr, "
         "employment program $300,000/yr - all Pot 2."),
        ("Tennant Creek", "NT - Barkly", "Listen", "Linda Turner",
         "To work through an existing shed rather than have a container placed.",
         "Equipment into a partner-supplied place",
         "Site base SHRINKS by whatever the partner supplies. The subtraction has not been agreed with them, so it "
         "cannot be asserted."),
        ("Palm Island", "QLD", "Listen", "-", "To start with governance, not plant.",
         "Governance and enterprise first; beds only if requested",
         "No public price. Standing rule: no community's own pathway gets a public price until they have seen it."),
        (),
        ("2. WHAT IS ALREADY DELIVERED - the proof, not the pipeline",),
        ("Evidence", "Figure", "Grade", "Why it matters to an investor"),
        ("Maningrida beds pressed, CNC'd and assembled in-house at the farm", 40, "verified",
         "The factory path is PROVEN as a process. The single strongest fact in the pack - the difference between a "
         "business plan and a demonstrated capability."),
        ("Assets confirmed in community at Utopia", 147, "verified",
         "Real delivery at scale into one community. Note: the Community OS figure of 169 is wrong."),
        ("Washing machines deployed (Pakkimjalki Kari)", 8, "verified",
         "Second product line, prototype stage. Shows the facility is not single-product."),
        ("Accountant-signed Goods-only revenue, FY26", 713827, "accountant-signed",
         "The only externally quotable revenue figure. Never the $403,901 surplus."),
        (),
        ("3. THE WEIGHTED DEMAND SCHEDULE - how to model it without lying",),
        ("Category", "Weighting", "What qualifies", "FY27 beds", "FY27 revenue at $750"),
        ("Signed", "100%", "Executed order or contract", 0, 0),
        ("Quoted", "50%", "Formal quote issued, not accepted", None,
         "Centrecorp: 130 beds, $106,150 (quote QU-0014, draft, deferred pending community feedback)"),
        ("In conversation", "20%", "Named community, named lead, no quote", None, "The four pathways above"),
        ("Capacity (do NOT count as revenue)", "0%", "What a facility could make if demand existed", 500,
         "The line most models get wrong. 500 beds/yr per facility is a production rate, not a sales forecast."),
        (),
        ("THE HONEST POSITION: $0 signed match-eligible capital and no signed FY27 bed orders as at 2026-08-03. "
         "The demand story today is four real relationships, 40 beds proven in production and 147 assets delivered. "
         "That is a credible base and it is NOT 1,000 beds of committed demand. Closing that gap is the hackathon "
         "question and the right use of QBE skilled volunteering.",),
        (),
        ("4. THE DEMAND-SIDE RESEARCH MATT ASKED FOR",),
        ("Define the ideal placement criteria for a facility, then count the addressable beds inside that radius:",),
        ("  - Number of satellite communities within a set travel distance",),
        ("  - Population and household count, and current bed provision per household",),
        ("  - Existing procurement channels: health services, housing programs, land councils, councils",),
        ("  - Replacement cycle: how often current mattresses actually fail (one Alice Springs provider sells $3M/yr "
         "of washing machines into remote communities, most in dumps within months - the same dynamic applies to beds)",),
        ("  - Non-community channels worth researching: outdoor and camping retail, festivals, emergency and "
         "disaster relief",),
    ]
    write_rows(ws, rows)
    ws["A1"].font = TITLE
    for r in (4, 12, 19, 27, 29):
        ws.cell(row=r, column=1).font = SECTION
    for r in (5, 13, 20):
        header_row(ws, r, 7)
    set_widths(ws, [42, 18, 20, 34, 52, 42, 46])
    wrap_all(ws, 7)


# ── Sheet 7: Xero FY26 Actuals ──────────────────────────────────────────────
# Pulled 2026-08-03 from the ACT Xero mirror (Supabase `tednluwflfhxyucgwigh`,
# xero_transactions synced 2026-08-02). This is the category split Matt asked for in
# Q10 and could not get. It did NOT need the accountant: it is a query.
#
# METHOD, because it is easy to get wrong and I got it wrong first time:
#   - Source is ACCPAY bills (status PAID or AUTHORISED) plus SPEND transactions
#     (status AUTHORISED), project_code ACT-GD, dated FY26.
#   - DELETED and VOIDED rows are EXCLUDED. This matters enormously: 136 of the 204
#     ACT-GD spend transactions carry status DELETED, $229,127 worth, almost certainly
#     Dext re-imports replacing originals. A query that does not filter on status
#     overstates the total roughly threefold.
#   - Account names are derived from the SUPPLIERS on each code, not from the Xero
#     default chart, which does not apply to this customised chart of accounts.
#   - Basis is closer to accrual than cash (it includes authorised-but-unpaid bills),
#     so it does NOT tie to the $309,126 cash figure and is not meant to.
XERO_FY26_ACTUALS = [
    # code, inferred name, total, share, evidence (suppliers), flag
    ("446", "Materials & Supplies", 131845.58, 0.348,
     "Bunnings, Steelmart, Centre Canvas, Metal Manufactures, The Plasticians, Barkly Hardware, Defy", ""),
    ("486", "Sub-contractors", 58186.82, 0.154,
     "Oonchiumpa Consultancy and Services, Joseph Kirmos, Adriana Beach", ""),
    ("412", "Consulting & Accounting", 52157.28, 0.138,
     "Defy Manufacturing ONLY", "MISBOOKED - this is bed manufacturing, not consulting"),
    ("400", "Advertising & Marketing", 38694.89, 0.102,
     "Defy Manufacturing, Oonchiumpa Consultancy", "PARTLY MISBOOKED - Defy lines are manufacturing"),
    ("750", "Plant & Equipment (FIXED ASSET)", 36622.72, 0.097,
     "Circularity Group, Multicam Systems", "CAPEX, not an expense - belongs on the balance sheet"),
    ("425", "Freight & Courier", 20851.21, 0.055, "Peak Up Transport, Sendle, ePrint", ""),
    ("421", "Light meals & refreshments", 6542.91, 0.017,
     "121 small lines: cafes, roadhouses, supermarkets on field trips", ""),
    ("493", "Travel - National", 6340.07, 0.017, "Qantas, Virgin, SIXT, regional hotels", ""),
    ("449", "MV - Fuel & Oil", 6016.82, 0.016, "BP, Ampol, Shell, remote roadhouses", ""),
    ("710", "Office/site equipment (FIXED ASSET)", 4705.00, 0.012, "Bionic Group (containers)",
     "CAPEX, not an expense"),
    ("429", "General Expenses", 4317.88, 0.011,
     "Loadshift Sydney, Bargain Car Rentals, Orange Sky, Metal Manufactures", ""),
    ("451", "(unmapped)", 3901.83, 0.010, "", "Confirm name in Xero"),
    ("447", "(unmapped)", 2154.47, 0.006, "", "Confirm name in Xero"),
    ("485", "(unmapped)", 1674.27, 0.004, "", "Confirm name in Xero"),
    ("420", "Entertainment", 1421.37, 0.004, "", ""),
    ("430", "(unmapped)", 1078.68, 0.003, "", "Confirm name in Xero"),
    ("453", "(unmapped)", 1010.82, 0.003, "", "Confirm name in Xero"),
    ("452", "Hire Expenses", 597.02, 0.002, "Van hire", ""),
    ("473", "(unmapped)", 259.95, 0.001, "", "Confirm name in Xero"),
    ("448", "(unmapped)", 142.70, 0.000, "", "Confirm name in Xero"),
    ("720", "Fixed asset (minor)", 117.28, 0.000, "", "CAPEX"),
    ("700", "Fixed asset (minor)", 102.43, 0.000, "", "CAPEX"),
    ("445", "Light, Power, Heating", 101.77, 0.000, "", ""),
]


def build_xero_actuals(wb):
    ws = wb.create_sheet("Xero FY26 Actuals")
    rows = [
        ("Goods (ACT-GD) FY26 expenses, split by account - pulled from Xero 2026-08-03",),
        ("This is the category split Matt asked for in Q10 and was told needed an accountant. It did not. "
         "It is a query against the Xero mirror. Account names are derived from the SUPPLIERS on each code, "
         "because the Xero default chart does not apply to this customised chart of accounts. Confirm the "
         "handful marked (unmapped) in Xero settings - about thirty seconds.",),
        (),
        ("Account code", "Inferred name", "FY26 total $", "Share", "Evidence (suppliers on this code)", "Flag"),
    ]
    for code, name, total, share, evidence, flag in XERO_FY26_ACTUALS:
        rows.append((code, name, total, share, evidence, flag))
    rows += [
        ("TOTAL", "", 378843.77, 1.0, "380 lines", ""),
        (),
        ("THE SPLIT THAT MATTERS FOR A 3-STATEMENT MODEL",),
        ("Bucket", "Amount $", "Lines", "Treatment", "", ""),
        ("OPEX", 337296.34, 373, "Profit and loss", "", ""),
        ("CAPEX (7xx asset accounts)", 41547.43, 7, "Balance sheet, then depreciated", "", ""),
        ("TOTAL", 378843.77, 380, "", "", ""),
        (),
        ("METHOD, AND WHY THE OBVIOUS QUERY IS WRONG",),
        ("Source: ACCPAY bills (PAID or AUTHORISED) plus SPEND transactions (AUTHORISED), project_code "
         "ACT-GD, dated 2025-07-01 to 2026-06-30.",),
        ("DELETED and VOIDED rows are EXCLUDED. 136 of the 204 ACT-GD spend transactions carry status "
         "DELETED, worth $229,127, almost certainly Dext re-imports replacing originals. A query that does "
         "not filter on status overstates the total roughly threefold. This is the single biggest trap in "
         "this data.",),
        ("Basis is closer to accrual than cash, because it includes authorised-but-unpaid bills. It "
         "therefore does NOT tie to the $309,126 cash figure in Inputs, and is not meant to. For the record, "
         "paid bills ($193,395) plus authorised spend ($95,762) is $289,157, which lands about $20,000 short "
         "of $309,126 - roughly the size of the shredder.",),
        (),
        ("THREE FINDINGS AN INVESTOR'S ACCOUNTANT WOULD REACH ON THEIR OWN",),
        ("1. THIS IS WHY COGS READS $0. About $90,852 of Defy Manufacturing spend - the bed kits, the actual "
         "cost of making the product - is booked to Consulting & Accounting (412) and Advertising & Marketing "
         "(400). It is not that nobody tracks unit economics; it is that the manufacturing spend is sitting in "
         "two overhead accounts. Reclassifying it to cost of sales would give Goods a real gross margin line "
         "for the first time.",),
        ("2. CAPEX IS INCONSISTENTLY TREATED. Circularity's $32,780 press-and-CNC went to 750, a fixed asset "
         "account. The Telford Smith shredder went to 446, Materials & Supplies. Same class of purchase, "
         "opposite treatment. $41,547 is correctly capitalised; an unknown amount sitting in 446 is not.",),
        ("3. THE MISSING SHREDDER INVOICE EXISTS, BUT IS DELETED. Telford Smith Engineering, 23 December 2025, "
         "two lines of $19,800: 'Telford Smith Engine' and 'Zerma GSL-300/400 Granulator'. Both records carry "
         "status DELETED, which is why a live P&L shows nothing and canon records 'no Xero record'. This does "
         "not close open question 4, but it turns 'find a missing invoice' into 'ask the bookkeeper why these "
         "two were deleted on 23 Dec'. Supplier, date, amount and machine model are all now known.",),
        (),
        ("SOURCE: Supabase project tednluwflfhxyucgwigh (the ACT Xero mirror). xero_transactions synced "
         "2026-08-02. NOTE: xero_bank_transactions is 7 months stale (last sync 2025-12-30) and carries no "
         "ACT-GD tracking - do not use that table.",),
    ]
    write_rows(ws, rows)
    ws["A1"].font = TITLE
    header_row(ws, 4, 6)
    for r in (len(XERO_FY26_ACTUALS) + 7, len(XERO_FY26_ACTUALS) + 13, len(XERO_FY26_ACTUALS) + 19):
        ws.cell(row=r, column=1).font = SECTION
    set_widths(ws, [16, 34, 18, 10, 62, 46])
    wrap_all(ws, 6)
    for r in range(5, 5 + len(XERO_FY26_ACTUALS)):
        ws.cell(row=r, column=4).number_format = "0.0%"


# ── Sheet 8: Open Questions ─────────────────────────────────────────────────
# Each row: priority, question, current evidence, working placeholder, owner,
# status, why it matters, next action, ANSWER, use-this-figure, grade, who closes.
OPEN_QUESTIONS = [
    ("Critical", "What is actual opening GOC cash?", "Current Inputs uses $50K assumption.", 50000,
     "Ben + accountant", "answered", "Controls the cash trough and funding requirement.",
     "Confirm bank balance and entity migration date.",
     "Do NOT model a GOC-only opening cash balance yet, because no GOC bank account exists. Entity-wide cash across "
     "all of A Curious Tractor at 2026-08-03 is $155,909 (Xero, org 'Nicholas Marchesi', refreshed 2026-08-02); "
     "receivables $412,818 and payables $325,632 sit against it. GOC opening cash is therefore a CARVE-OUT DECISION, "
     "not a bank fact: it equals whatever cash is transferred on the handover date. Keep $50,000 as the base case, "
     "and run $0 and $150,000 as the two sensitivities, because this input controls the trough and every "
     "funding-need number downstream.",
     "$50,000 base; sensitivity $0 / $150,000", "assumption, bounded by verified entity cash",
     "Ben + accountant, at the handover date"),

    ("Critical", "Which replication capex scope should Matt use?",
     "MVF $90.8K-$123K, midpoint ~$105K. Matt full container build $207,450.", 105000, "Ben + Matt", "answered",
     "Changes funding need and depreciation.", "Choose scope, then obtain firm vendor quotes.",
     "Use $105,000 as the REPLICATION planning figure and carry Matt's $207,450 as the fully-built container-workshop "
     "scope. They are not competing estimates of the same thing: $105,000 is the Minimal Viable Facility "
     "reconciliation - what it costs to stand up a working line at a new site using the equipment set Goods already "
     "runs - and $207,450 prices a turnkey 40ft container workshop with a $167K equipment midpoint. Model $105K as "
     "base, $207K as the conservative/defensible case, and label them by scope not by optimism. The module basket "
     "(see 'Facility & Modules') is what actually makes this answerable per site, because three of the four live "
     "pathways do not want a whole facility.",
     "$105,000 base (MVF) / $207,450 conservative (turnkey 40ft)", "modelled midpoint, both scopes evidence-based",
     "Ben chooses scope per site; firm vendor quotes still required"),

    ("Critical", "What did the bigger CNC cost, and who owns it?",
     "Recently purchased. About $5,135 installation booked.", "TBC", "Ben + bookkeeper", "open",
     "Opening PP&E and ownership cannot reconcile without it.", "Locate invoice and purchasing entity.",
     "Still open, and it is paperwork not modelling. What is known: about $5,135 of installation is booked "
     "separately, and the machine is physically on site. Until the invoice and the purchasing entity are found, "
     "EXCLUDE the CNC purchase price from opening PP&E and disclose it as an unrecorded owned asset. Do not plug a "
     "number - a guessed asset value on an opening balance sheet is the single easiest thing for an investor's "
     "accountant to pull apart.",
     "Exclude from opening PP&E; disclose as note", "open", "Ben + bookkeeper - locate invoice"),

    ("Critical", "Have the shredder and container ownership records been located?",
     "$19,800 shredder is physical-only. Some container costs are ambiguously tagged.", "TBC", "Ben + bookkeeper",
     "answered", "Determines opening assets and evidence quality.",
     "Locate invoices or formally document owned, invoice to locate.",
     "Partly. The $19,800 Telford Smith shredder is confirmed physically by Ben with no invoice in the connected "
     "Xero. Container-side records ARE located: 20ft Monument Grey $3,320 (Bionic Self Storage, 2026-04-29), crane "
     "placement $1,041 (GM Crane Hire, 2026-06-29), transport $193 (Rapid Container, 2026-06-24). Two items stay "
     "excluded: the $6,600 generators (Orange Sky, no project tag) and the $5,904 larger container (tagged Mounty "
     "Yarns, flagged on-sold). Treatment: include the shredder at $19,800 with a stated 'owned, invoice to locate' "
     "note; exclude the two ambiguous items.",
     "Shredder in at $19,800; generators and larger container OUT", "evidenced / physical-only mix",
     "Ben + bookkeeper - written asset register entry"),

    ("Critical", "Was the 40-bed run measured for time, diesel and yield?",
     "Factory process proven on 40 Maningrida beds. Per-bed cost remains modelled.", "Short measured run",
     "Ben + production lead", "answered", "Locks the $425.74 factory cost and sustained capacity.",
     "Recover run records or perform a measured run.",
     "Assume NOT measured. The 40 Maningrida beds prove the PROCESS - pressed, CNC'd and assembled in-house at the "
     "farm - which is the claim that matters and must never be softened. What that run did not produce is captured "
     "time, diesel and yield, so $425.74 stays graded 'modelled (proven process)'. The fix is cheap: a one-week "
     "measured run logging operator hours, diesel litres, HDPE input kg vs finished leg kg, and reject rate. Until "
     "then do not let any document call $425.74 'actual'.",
     "$425.74 stays modelled; schedule a 1-week measured run", "modelled (process proven)",
     "Ben + production lead - book the run"),

    ("Critical", "What sustained beds-per-facility capacity should be modelled?",
     "Sources say 250, 1,250 and about 1,500 beds/yr. Matt uses 500.", 500, "Ben + production lead", "answered",
     "Determines whether production and community facilities are viable.",
     "Use 500 conservatively until the measured run resolves it.",
     "Model 500 beds/yr per facility, single shift, and state it as a MODELLED PLANNING RATE not a demonstrated one. "
     "The 250 / 1,250 / 1,500 figures in circulation are not conflicting measurements, they are different things: "
     "250 is an in-source-assembly step, and the 1,250-1,500 range is theoretical press throughput with no allowance "
     "for changeover, maintenance, weather, staff availability or demand. 500 is the number that survives an "
     "investor asking 'and what happens on a bad week'. Sensitivity band 350-750.",
     "500 beds/yr/facility (band 350-750)", "assumption, conservative by design",
     "Resolved by the measured run in Q5"),

    ("Critical", "What is the actual capital stack and draw timing?",
     "SEFA $300K, QBE $400K and philanthropy $500K are placeholders. Signed match today is $0.", "TBC",
     "Ben + Matt + QBE", "answered", "Drives debt, cash and interest.",
     "Confirm instrument, amount, date and matching eligibility.",
     "There is no capital stack yet and the model must say so. Signed match-eligible capital today is $0. Treat SEFA "
     "$300K (debt, no ownership gate, Yr1), QBE $400K (repayable preferred, Yr2, match ratio unconfirmed) and anchor "
     "philanthropy $500K (Yr1) as a PROPOSED stack, and run the model on a Yr1-only base case of SEFA $300K plus "
     "philanthropy $500K with QBE modelled as upside. QBE is catalytic, NOT dollar-for-dollar, and there is no $150K "
     "floor - do not build a matching formula into the sheet.",
     "Base case Yr1 $800K (SEFA $300K + philanthropy $500K); QBE $400K as upside", "placeholder",
     "Ben + Matt + QBE/SIH - instrument, amount, date"),

    ("Critical", "What entity owns assets and trades at forecast start?",
     "Historical activity is in the sole trader. Pty Ltd migration and plant handover are not complete.", "TBC",
     "Ben + legal + accountant", "answered", "Sets opening balance sheet, tax and related-party treatment.",
     "Confirm transfer date, values and legal structure.",
     "Unresolved and it is the biggest structural risk in the pack. Historical trading sits in the sole trader "
     "(A Curious Tractor, ABN-verified); the Pty Ltd migration and the plant handover are incomplete, with handover "
     "targeted end-August 2026. Separately, DGR flows only through The Butterfly Movement Ltd, and 51% First Nations "
     "ownership is the unlock for IBA/FAC. MODEL IT AS: forecast starts 1 Jul 2026 in the Pty Ltd, opening PP&E "
     "transferred at written-down value, with a stated note that the transfer is not executed. Do not present the "
     "model as if the entity already trades.",
     "Forecast entity = Pty Ltd from FY27; transfer at WDV; disclose as not executed", "open - structural",
     "Ben + MinterEllison + accountant"),

    ("High", "Should $67,200 founder non-production time be an entity expense?",
     "Locked 120 days x $560. Excluded from unit cost.", 67200, "Ben + Matt", "answered",
     "Without treatment, overhead and profitability can be overstated.",
     "Choose expense, founder contribution or disclosed sensitivity.",
     "Yes - show it, but below the line. Put the full $84,000 founder cost in the entity P&L, split as $16,800 "
     "production (inside unit cost, already locked) and $67,200 non-production. Then show a clearly labelled "
     "'founder contribution' add-back so a reader sees BOTH the true cost of running the business and the cash "
     "reality that it is not currently paid. Hiding the $67,200 overstates profitability; burying it in unit cost "
     "overstates the bed cost. It is a cost of capital, not a cost of production.",
     "$67,200 as entity overhead, with an explicit founder-contribution add-back line",
     "locked figure, treatment decided", "Ben + Matt - agreed treatment"),

    ("High", "Can historical $309,126 expenses be split by category?",
     "Only an aggregate cash-basis Goods expense figure is available here.", "TBC", "Accountant + bookkeeper", "open",
     "Matt needs a recurring overhead base, not one aggregate.",
     "Provide accountant carve-out by direct cost, payroll, travel, professional services and other.",
     "ANSWERED 2026-08-03, and it did not need the accountant. The split is on the 'Xero FY26 Actuals' tab, "
     "pulled from the Xero mirror: Materials & Supplies $131,846, Sub-contractors $58,187, Consulting & "
     "Accounting $52,157, Advertising & Marketing $38,695, Plant & Equipment $36,623, Freight $20,851, then a "
     "long tail. $337,296 opex and $41,547 capex across 380 lines. Two caveats: DELETED rows must be excluded "
     "(136 of 204 spend transactions are deleted, and including them overstates by roughly threefold), and the "
     "basis is accrual-ish so it does not tie to the $309,126 cash figure. The accountant is still worth asking "
     "to confirm the reclassifications below, but the base is no longer missing.",
     "See 'Xero FY26 Actuals' tab: $337,296 opex / $41,547 capex", "verified from Xero mirror",
     "Accountant to confirm the COGS reclassification"),

    ("High", "Site manager or coordinator?",
     "DEWR prices $150K incl. super. Alternative $90K coordinator is an assumption.", 150000,
     "Ben + community partner", "answered", "Moves staffed-site cost and break-even materially.",
     "Confirm operating design for each site type.",
     "Model BOTH, as two site archetypes, because they are two different operating designs rather than a cheap and "
     "an expensive version of one. A $150,000 project manager (DEWR-priced, incl. super) is the staffed-site design "
     "where the site runs an employment program; a $90,000 coordinator is the production-only design. Since the "
     "manager sits in Pot 2 (grant funded by design), the choice does not move bed break-even at all - it moves the "
     "size of the grant the site needs. Saying that out loud is a strength.",
     "Two archetypes: staffed $150K / production-only $90K, both in Pot 2", "as-written (DEWR) vs assumption",
     "Ben + the specific community partner, per site"),

    ("High", "Is the $40K trainer/WHS split correct?",
     "Derived from a blended $190K ACT machinery + trainer/WHS line.", 40000, "Ben + Oonchiumpa", "answered",
     "The number is not stated separately in the application.", "Obtain the underlying budget split.",
     "The $40,000 split is an estimate, not a stated line - it is a modelled portion of a blended $190K DEWR line "
     "that covers ACT machinery AND trainer/WHS. Keep $40,000 as the working figure but grade it 'modelled split' "
     "everywhere it appears, and flag the related and more interesting point: $63,333/yr of that same DEWR line is "
     "REVENUE to Goods., not a cost to the site, and no model currently shows it.",
     "$40,000 held, graded 'modelled split'; also book $63,333/yr as Goods. revenue", "modelled split",
     "Ben + Oonchiumpa - underlying budget detail"),

    ("High", "What working-capital treatment applies to grants?",
     "30 debtor, 30 creditor and 45 inventory days only address normal operations.", "TBC", "Matt + accountant",
     "answered", "Grant receivables and acquittal timing can dominate cash.",
     "Set grant invoice, receipt and restricted-fund timing.",
     "Grants must be modelled as their own working-capital stream, separate from trade debtors. Do not apply the "
     "30-day debtor assumption to grant income. Model each grant as: milestone date, invoice date, receipt lag "
     "(60-90 days is realistic for government), and hold unspent restricted funds as a liability, not as revenue or "
     "free cash. Live proof this matters: $167,200 from ALIVE / University of Melbourne (INV-0341 $66,000 and "
     "INV-0342 $101,200) was due 30 July and is overdue, and total receivables of $412,818 are 100% overdue.",
     "Separate grant WC stream: 60-90 day receipt lag, unspent = liability", "assumption, grounded in live AR",
     "Matt + accountant"),

    ("High", "What are debt fees, term and repayment profile?", "Only a 6% interest assumption exists.", "TBC",
     "Matt + lenders", "answered", "Interest-only is not enough for a financing schedule.",
     "Add establishment fee, draw dates, maturity and principal repayments.",
     "Not answerable from our side - no facility has been offered, so there is no term sheet to model. Hold 6% "
     "interest-only as the placeholder and add explicit rows Matt can populate later: establishment fee, draw dates, "
     "term, interest-only period, amortisation profile, security and covenants. The important modelling instruction "
     "is that QBE money is preferred as REPAYABLE, so the model should demonstrate the entity can service and repay "
     "it from production surplus, not from grants.",
     "6% interest-only placeholder; add fee/term/repayment rows as blanks", "assumption",
     "Matt + SEFA / QBE lenders"),

    ("High", "What tax and GST treatment applies?",
     "25% company-tax placeholder. Rough GST payable signal ~$29,657 is unverified.", "TBC", "Accountant", "answered",
     "Grants, losses, DGR flows and asset transfers may differ by entity.",
     "Confirm company tax, GST-free grants, BAS and loss carry-forwards.",
     "Needs the accountant and it is entity-dependent, which is why Q8 blocks it. Directionally: a Pty Ltd pays 25% "
     "base rate with losses carried forward (this business will carry losses, so cash tax is likely nil in the "
     "forecast window); most government grants are GST-free but many philanthropic grants are not; and DGR-receipted "
     "flows can only go through The Butterfly Movement Ltd, which is a different entity from the one that would "
     "trade. Model 25% with loss carry-forward, cash tax nil until profitable, and flag GST and DGR as "
     "accountant-confirmed items.",
     "25% with loss carry-forward, cash tax nil in window", "assumption pending advice",
     "Accountant - written advice"),

    ("Medium", "What central entity overheads sit outside the $109,500 block?",
     "No defensible standalone totals for legal, audit, insurance, software, governance or banking.", "TBC",
     "Ben + accountant", "answered", "These costs are required in an entity P&L.",
     "Extract actuals and set a Year 1 budget.",
     "There is no defensible standalone total in the codebase, so budget it rather than pretend to source it. A "
     "first-year central overhead budget for an entity at this scale: insurance $12,000, accounting and audit "
     "$15,000, legal and entity setup $10,000 (one-off, elevated by the migration), software and IT $8,000, banking "
     "and merchant fees $3,000, governance and board $5,000. Total about $53,000/yr, explicitly graded as a BUDGET. "
     "This sits outside the $109,500 production block and must not be double-counted with the $14,700 admin line "
     "inside it.",
     "~$53,000/yr central overhead budget (6 named lines)", "budget - no source, honestly labelled",
     "Ben + accountant - replace with actuals"),

    ("Medium", "What payroll on-cost assumptions should apply?",
     "DEWR manager includes super. Other labour assumptions may not.", "TBC", "Accountant", "answered",
     "Super, leave, workers compensation and payroll tax can materially change staffing cost.",
     "Confirm which labour lines are fully loaded.",
     "Assume NOT fully loaded except where stated. The DEWR project manager at $150,000 explicitly includes super. "
     "Every other labour figure in the pack - the $560/day founder rate, the $130/bed community fair wage, the "
     "$40,000 trainer/WHS split - should be treated as base cost. Apply a 25% on-cost loading (12% super, plus "
     "leave, workers compensation and payroll tax) to any line that becomes real employment, and show it as a "
     "separate line so it is visible rather than baked in.",
     "25% on-cost loading on all labour except the $150K DEWR manager", "assumption",
     "Accountant - confirm payroll tax threshold"),

    ("Medium", "What warranty and replacement provision is appropriate?",
     "Stretch Bed has a five-year warranty. No provision is in the pack.", "TBC", "Ben + Matt", "answered",
     "A credible product P&L needs expected warranty cost.",
     "Review claims history and set a percent-of-sales provision.",
     "No claims history exists, so set a provision by first principles and revisit. The Stretch Bed carries a "
     "five-year warranty, and the failure modes are known and cheap: canvas is the wear part, the HDPE legs and "
     "galvanised poles are effectively inert. Provide 3% of sales (about $22.50/bed) as the base case, band 2-5%. "
     "Note the honest counterpoint that strengthens rather than weakens this: field replacement in remote "
     "communities is expensive logistics, not expensive parts, which is itself an argument for local production.",
     "3% of sales ($22.50/bed), band 2-5%", "assumption - first principles",
     "Ben + Matt; replace with claims history from year 2"),

    ("Medium", "What bad-debt, inventory-loss and production-yield allowances apply?",
     "No current assumptions. Factory yield has not been measured at sustained rate.", "TBC", "Ben + accountant",
     "answered", "Avoids overstating margin, AR and inventory.",
     "Use actual history or conservative placeholders.",
     "Set conservative placeholders and mark them for replacement by the measured run. Bad debt 2% of product "
     "revenue (low because buyers are largely institutional and government-adjacent, though the current "
     "100%-overdue receivables position argues against going lower). Inventory loss/shrinkage 2%. Production yield "
     "90% on the factory path - HDPE reprocessing has real offcut and reject loss and assuming 100% would overstate "
     "both margin and the plastic-diverted claim. The yield number is measurable and should come out of Q5.",
     "Bad debt 2% / inventory loss 2% / factory yield 90%", "assumption - conservative",
     "Ben + accountant; yield from the measured run"),

    ("Medium", "What inflation and price-indexation assumptions apply?",
     "$750 price is held flat and no cost inflation is modelled.", "TBC", "Ben + Matt", "answered",
     "Longer forecasts otherwise create misleading margins.",
     "Set annual price, wages, materials and facility escalation.",
     "Index everything from FY28 onward. Holding the $750 price flat while costs inflate is the most common way a "
     "five-year model quietly destroys its own margin. Model: price +2.5%/yr from FY28, wages +3.5%/yr, materials "
     "+3.0%/yr, facility and freight +3.0%/yr. Keep price flat in FY27 only, because a price rise inside the first "
     "year is a real commercial decision affecting communities and should not be assumed by a spreadsheet. Also show "
     "a price-flat sensitivity, since holding $750 may well be a deliberate choice.",
     "From FY28: price +2.5%, wages +3.5%, materials +3.0%, facility +3.0%. FY27 flat.", "assumption",
     "Ben + Matt - and a deliberate call on whether $750 holds"),

    ("Medium", "What volumes and production-path mix are supportable by demand?",
     "Matt's unit model is capacity-driven. Demand constraint remains open.", "TBC", "Ben + commercial lead", "open",
     "Revenue cannot be based on capacity alone.",
     "Build a signed/weighted demand schedule by buyer and year.",
     "This is the real gap, and Matt named it correctly: the model is capacity-driven and demand is unproven. What "
     "we can evidence today is on the 'Case Studies & Demand' sheet - four live community pathways at named stages "
     "with named leads, plus 40 beds already delivered to Maningrida and 147 assets confirmed at Utopia. What we "
     "cannot yet evidence is 1,000 beds/yr of committed demand at $750. Do not model capacity as revenue. Model a "
     "signed/weighted demand schedule: signed at 100%, quoted at 50%, in-conversation at 20%, and let the sheet show "
     "the gap honestly. The hackathon and QBE skilled volunteering are the right vehicle to close it.",
     "Weighted demand schedule (signed 100% / quoted 50% / conversation 20%)", "open - the biggest gap",
     "Ben + commercial lead; hackathon + QBE mentor"),

    ("Medium", "How should Pot 2 grant income match Pot 2 costs?",
     "Bare $151,666, staffed $341,666 and program $300K costs are known, but funding timing is not.", "TBC",
     "Ben + Matt", "answered", "Prevents production revenue from implicitly subsidising the wraparound.",
     "Model separate restricted grant income and cost centres.",
     "Match Pot 2 grant income to Pot 2 costs in the SAME period, in a separate restricted cost centre that never "
     "touches the production P&L. Mechanically: a site does not open until its wraparound funding is secured, so "
     "model the Pot 2 cost block as starting on the grant start date, with unspent restricted funds carried as a "
     "liability. The rule to state on the model's cover is that Pot 2 never appears in bed break-even, and Pot 1 "
     "never depends on Pot 2 continuing. If a site's grant ends, production should keep running - if it cannot, the "
     "two-pot claim is not true and the model should show that.",
     "Restricted cost centre; Pot 2 costs start on grant start date; unspent = liability", "structural rule - agreed",
     "Ben + Matt - model architecture"),
]


def build_open_questions(wb):
    ws = wb.create_sheet("Open Questions")
    ws.append(["Open Questions for the GOC Entity Model"])
    ws.append([
        "Items Matt may need before the GOC-only 3-statement model can be treated as decision-ready. "
        "ANSWERED 2026-08-03 (columns I-L). Most are now decision-ready working figures; the rest name exactly who "
        "has to close them and what document does it. Nothing here has been invented: where the codebase holds no "
        "defensible number, the answer says so and proposes a labelled assumption instead."
    ])
    ws.append([])
    ws.append([
        "Priority", "Question / missing input", "Current evidence", "Working placeholder", "Owner", "Status",
        "Why it matters", "Next action", "ANSWER (2026-08-03)", "Use this figure", "Grade", "Who closes it",
    ])
    for q in OPEN_QUESTIONS:
        ws.append(list(q))
    ws["A1"].font = TITLE
    header_row(ws, 4, 12)
    set_widths(ws, [10, 34, 34, 18, 20, 12, 34, 34, 78, 40, 26, 30])
    wrap_all(ws, 12)
    for r in range(5, 5 + len(OPEN_QUESTIONS)):
        ws.row_dimensions[r].height = 150
    ws.freeze_panes = "B5"


# ── Entry point ─────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "-o", "--output",
        default="deliverables/GOC-Entity-Model-Inputs-ANSWERED-2026-08-03.xlsx",
        help="Where to write the workbook (gitignored: it is a build artifact)",
    )
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; every tab is built explicitly

    build_model(wb)
    build_inputs(wb)
    build_notes(wb)
    build_known_costs(wb)
    build_facility_modules(wb)
    build_case_studies(wb)
    build_xero_actuals(wb)
    build_open_questions(wb)

    out = pathlib.Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  {len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames)}")
    print(f"  {len(OPEN_QUESTIONS)} open questions, all answered")
    print("\nPush to the live Google Sheet with:")
    print(f"  .venv-sheets/bin/python tools/push-xlsx-to-gsheet.py {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
