#!/usr/bin/env python
"""
Push a local .xlsx into a Google Sheet, tab for tab, formulas intact.

Why this exists: the MCP Drive upload needs the whole file inlined as base64,
which is not something an agent can transcribe reliably. This reads the workbook
locally and writes cell values through the Sheets API instead, so the round trip
is exact and repeatable. Run it again after editing the workbook and it updates
the SAME sheet rather than making a new one.

Auth, once. Three routes are tried in order, because two of the obvious ones are
commonly blocked:

  1. Your own OAuth client ID at ~/.config/goods-sheets-oauth.json (PREFERRED).
     Google now blocks Drive/Sheets scopes on gcloud's default client ID, and many
     orgs enforce `iam.disableServiceAccountKeyCreation`, which kills route 2. A
     Desktop-app client ID of your own is subject to neither, and the resulting
     Sheet is owned by you rather than by a service account.
  2. A service account key at ~/.config/goods-sheets-sa.json.
  3. gcloud application-default credentials, if they ever work again.

Usage:
    .venv-sheets/bin/python tools/push-xlsx-to-gsheet.py <file.xlsx> [--title "Sheet name"]
    .venv-sheets/bin/python tools/push-xlsx-to-gsheet.py <file.xlsx> --key <existing_sheet_id>

Formulas are written with USER_ENTERED so Sheets parses "=SUM(...)" as a formula,
not as text. openpyxl is opened with data_only=False for the same reason: we want
the formula string, never the last cached value Excel happened to leave behind.
"""
import argparse
import json
import pathlib
import sys

import gspread
import openpyxl
from google.auth import default as google_auth_default

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Where we remember which Google Sheet a given workbook maps to, so a second run
# updates in place instead of littering the Drive with near-identical copies.
STATE = pathlib.Path(__file__).parent / ".gsheet-targets.json"

CONFIG = pathlib.Path.home() / ".config"
OAUTH_CLIENT = CONFIG / "goods-sheets-oauth.json"      # your own Desktop client ID
OAUTH_TOKEN = CONFIG / "goods-sheets-token.json"       # cached after the first login
SERVICE_ACCOUNT = CONFIG / "goods-sheets-sa.json"


def authorize() -> gspread.Client:
    """Return an authorised gspread client, trying the routes most likely to work first."""
    if OAUTH_CLIENT.exists():
        # Opens a browser once, then reuses the cached token forever.
        return gspread.oauth(
            credentials_filename=str(OAUTH_CLIENT),
            authorized_user_filename=str(OAUTH_TOKEN),
            scopes=SCOPES,
        )

    if SERVICE_ACCOUNT.exists():
        return gspread.service_account(filename=str(SERVICE_ACCOUNT), scopes=SCOPES)

    creds, _ = google_auth_default(scopes=SCOPES)
    return gspread.authorize(creds)


def load_state() -> dict:
    if STATE.exists():
        return json.loads(STATE.read_text())
    return {}


def save_state(state: dict) -> None:
    STATE.write_text(json.dumps(state, indent=2) + "\n")


def sheet_rows(ws) -> list[list]:
    """Worksheet -> list of row lists, formulas as strings, None -> ''."""
    rows = []
    for row in ws.iter_rows():
        rows.append(["" if c.value is None else c.value for c in row])
    # Trim wholly empty trailing rows so we do not pad the Google Sheet with blanks.
    while rows and not any(str(v).strip() for v in rows[-1]):
        rows.pop()
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--title", help="Title for a newly created sheet")
    ap.add_argument("--key", help="Push into this existing spreadsheet id")
    ap.add_argument("--share", help="Email to give writer access to")
    args = ap.parse_args()

    path = pathlib.Path(args.xlsx).expanduser().resolve()
    if not path.exists():
        print(f"No such workbook: {path}", file=sys.stderr)
        return 1

    gc = authorize()

    state = load_state()
    key = args.key or state.get(str(path))

    if key:
        sh = gc.open_by_key(key)
        print(f"Updating existing sheet: {sh.title}")
    else:
        title = args.title or path.stem
        sh = gc.create(title)
        print(f"Created sheet: {title}")

    wb = openpyxl.load_workbook(path, data_only=False)

    # Google Sheets refuses to delete the last remaining worksheet, so we build the
    # new tabs first and only then remove whatever was there before.
    pre_existing = {ws.title: ws for ws in sh.worksheets()}
    written = []

    for name in wb.sheetnames:
        rows = sheet_rows(wb[name])
        n_rows = max(len(rows), 1)
        n_cols = max((len(r) for r in rows), default=1)

        target_name = name if name not in pre_existing else f"{name} (new)"
        ws = sh.add_worksheet(title=target_name, rows=n_rows + 20, cols=n_cols + 5)
        if rows:
            ws.update(rows, "A1", value_input_option="USER_ENTERED")
        ws.freeze(rows=1)
        written.append((ws, name))
        print(f"  wrote {name}: {n_rows} rows x {n_cols} cols")

    for ws in pre_existing.values():
        sh.del_worksheet(ws)

    # Restore the intended tab names now the originals are gone.
    for ws, intended in written:
        if ws.title != intended:
            ws.update_title(intended)

    # Match the workbook's tab order.
    sh.reorder_worksheets([ws for ws, _ in written])

    if args.share:
        sh.share(args.share, perm_type="user", role="writer")
        print(f"  shared with {args.share}")

    state[str(path)] = sh.id
    save_state(state)

    print(f"\nDone: {sh.url}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
